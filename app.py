import os, io, uuid, re, base64, json
from datetime import datetime, date
from flask import Flask, request, jsonify, render_template, send_file, Response
try:
    from supabase import create_client, Client as SupabaseClient
    _DB_AVAILABLE = True
except ImportError:
    _DB_AVAILABLE = False
import fitz  # PyMuPDF
import pandas as pd
import openpyxl
import anthropic

# ── Load .env file if present ──────────────────────────────────────────────────
_ENV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.isfile(_ENV_PATH):
    with open(_ENV_PATH) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _k, _v = _line.split('=', 1)
                _k = _k.strip(); _v = _v.strip().strip('"').strip("'")
                # Set if the env var is missing or empty, and the file value is non-empty / non-placeholder
                if _v and _v != 'your-api-key-here' and not os.environ.get(_k):
                    os.environ[_k] = _v
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ROOT = os.path.dirname(os.path.abspath(__file__))
BRAND_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Style', 'Powered by Wellx Labs no background.png')
app = Flask(__name__,
            root_path=_ROOT,
            template_folder=os.path.join(_ROOT, 'templates'),
            instance_path=os.path.join(_ROOT, 'instance'))
app.secret_key = os.urandom(24)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB

# ── Session helpers (Supabase-backed, survives serverless cold starts) ─────────
import base64 as _b64
from datetime import timedelta

def _session_set(token: str, data: dict):
    """Persist session to Supabase. bytes values are base64-encoded."""
    def _enc(v):
        if isinstance(v, (bytes, bytearray)):
            return {'__b64__': True, 'data': _b64.b64encode(bytes(v)).decode()}
        return v
    serialised = {k: _enc(v) for k, v in data.items()}
    try:
        # Cleanup sessions older than 6 hours while we're here
        cutoff = (datetime.utcnow() - timedelta(hours=6)).isoformat()
        supa.table('sessions').delete().lt('created_at', cutoff).execute()
    except Exception:
        pass
    supa.table('sessions').upsert({'token': token, 'data': serialised}).execute()

def _session_get(token: str):
    """Retrieve session from Supabase. Decodes base64 binary values."""
    try:
        res = supa.table('sessions').select('data').eq('token', token).limit(1).execute()
    except Exception:
        return None
    if not res.data:
        return None
    raw = res.data[0]['data']
    def _dec(v):
        if isinstance(v, dict) and v.get('__b64__'):
            return _b64.b64decode(v['data'])
        return v
    return {k: _dec(v) for k, v in raw.items()}

def _session_patch(token: str, updates: dict):
    """Update specific keys in an existing session without overwriting the rest."""
    stored = _session_get(token)
    if stored is None:
        return
    stored.update(updates)
    _session_set(token, stored)

# ── Constants ────────────────────────────────────────────────────────────────
BASMAH_FEE       = 37.0
VAT_RATE         = 0.05
MAT_AGE_MIN      = 18
MAT_AGE_MAX_DXB  = 45
MAT_AGE_MAX_AUH  = 50

DEFAULT_BRACKETS = [
    (0,10,"0-10"), (11,17,"11-17"), (18,25,"18-25"),
    (26,30,"26-30"), (31,35,"31-35"), (36,40,"36-40"),
    (41,45,"41-45"), (46,50,"46-50"), (51,55,"51-55"),
    (56,59,"56-59"), (60,64,"60-64"), (65,99,"65-99"),
]

DOH_BRACKETS = [
    (0, 17, "0-17"), (18, 40, "18-40"),
    (41, 60, "41-60"), (61, 99, "61-99"),
]

# ── PDF Utilities ─────────────────────────────────────────────────────────────
def pdf_page_image(pdf_bytes, page_idx, scale=2.0):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    if page_idx < 0 or page_idx >= len(doc):
        page_idx = len(doc) - 1
    pix = doc[page_idx].get_pixmap(matrix=fitz.Matrix(scale, scale))
    return base64.b64encode(pix.tobytes("png")).decode(), len(doc)

def try_extract_text(pdf_bytes):
    try:
        import pdfplumber
        pages = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for pg in pdf.pages:
                pages.append(pg.extract_text() or "")
        return pages
    except Exception:
        return []

def find_rate_page_idx(page_texts, pdf_bytes):
    keywords = ["PREMIUM SUMMARY", "Age Range", "Category A", "PREMIUM OVERVIEW"]
    for i, text in enumerate(page_texts):
        if text and any(kw.lower() in text.lower() for kw in keywords):
            return i
    # Try pytesseract OCR fallback
    try:
        import pytesseract
        from PIL import Image
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for i in range(len(doc)):
            pix = doc[i].get_pixmap(matrix=fitz.Matrix(2.5, 2.5))
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            text = pytesseract.image_to_string(img)
            if any(kw.lower() in text.lower() for kw in keywords):
                return i
    except Exception:
        pass
    # Heuristic: rate table is often in the last quarter of the document
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    n = len(doc)
    return max(0, n - 4)  # Try 4th-from-last page

def try_parse_rates_from_text(text):
    rates = []
    pattern = re.compile(
        r'\b(\d{1,2})\s*[-–]\s*(\d{2,3})\b\s+([\d,]{4,})\s*(?:\d+\s+)?([\d,]{4,})?',
        re.MULTILINE
    )
    seen = set()
    for m in pattern.finditer(text):
        lo, hi = int(m.group(1)), int(m.group(2))
        if lo > hi or hi > 120:
            continue
        male_rate = int(m.group(3).replace(',', ''))
        if male_rate < 100:
            continue
        fem_str = m.group(4) or m.group(3)
        fem_rate = int(fem_str.replace(',', ''))
        if fem_rate < 100:
            fem_rate = male_rate
        key = (lo, hi)
        if key not in seen:
            seen.add(key)
            rates.append({'age_lo': lo, 'age_hi': hi, 'label': f"{lo}-{hi}",
                          'male': male_rate, 'female': fem_rate})
    return sorted(rates, key=lambda x: x['age_lo'])

def try_parse_maternity_rate(text):
    m = re.search(r'Additional\s+Maternity\s+Premium\s+(?:AED\s+)?([\d,]+)', text, re.I)
    if m:
        v = int(m.group(1).replace(',', ''))
        return v if v > 100 else 0
    return 0

def extract_rates_with_claude_vision(pdf_bytes, page_idx):
    """Use Claude Vision API to extract premium rates from an image-based PDF page."""
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        print("[Vision] ANTHROPIC_API_KEY not set — skipping Claude Vision extraction")
        return None, {}

    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        if page_idx < 0 or page_idx >= len(doc):
            page_idx = len(doc) - 1
        pix = doc[page_idx].get_pixmap(matrix=fitz.Matrix(3, 3))
        img_b64 = base64.standard_b64encode(pix.tobytes("png")).decode()
        doc.close()
    except Exception as e:
        print(f"[Vision] Failed to render PDF page: {e}")
        return None, {}

    prompt = """You are extracting premium rate tables from an insurance quote PDF page.

Return ONLY a valid JSON object — no markdown, no explanation — with this exact structure:
{
  "categories": {
    "<letter>": {
      "brackets": [
        {"label": "<lo>-<hi>", "age_lo": <int>, "age_hi": <int>, "male": <float>, "female": <float>}
      ],
      "maternity_rate": <float or 0>
    }
  },
  "quote_totals": {
    "total_premium": <float or 0>,
    "members": <int or 0>,
    "grand_total": <float or 0>
  }
}

Rules:
- Extract ALL categories visible (A, B, C, etc.). If only one category, use "A".
- "male" and "female" are annual premiums in AED. If the table shows one rate column, use the same value for both.
- "maternity_rate" is the PER-ELIGIBLE-FEMALE maternity surcharge. Look for a line labelled "Additional Maternity Premium" or "Additional Maternity Premium / Married Females". The FIRST value shown for that label is the per-female rate (e.g. AED 5,264). A separate line may show "Total Maternity Premium" (e.g. AED 21,055) — that is the SUM across all eligible females and must NOT be used. Extract ONLY the per-female rate. Use 0 if the label is completely absent.
- "total_premium" is the net premium before fees (look for "TOTAL PREMIUM"). Use 0 if not shown.
- "members" is the total member count from the summary. Use 0 if not shown.
- "grand_total" includes BASMAH + VAT. Use 0 if not shown.
- Age bracket labels must match the lo-hi format exactly, e.g. "0-10", "11-17", "18-25"."""

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=2048,
            system="You are a precise structured data extractor for insurance documents. Return only valid JSON.",
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": img_b64,
                        },
                    },
                    {"type": "text", "text": prompt}
                ],
            }]
        )
        raw = response.content[0].text.strip()
        # Strip markdown fences if Claude wraps the JSON
        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)
        data = json.loads(raw)
        cats = data.get('categories', {})
        totals = data.get('quote_totals', {})
        print(f"[Vision] Extracted {len(cats)} categorie(s): {list(cats.keys())}")
        return cats, totals
    except Exception as e:
        print(f"[Vision] Claude Vision extraction failed: {e}")
        return None, {}

def try_parse_quote_totals(text):
    totals = {}
    m = re.search(r'TOTAL\s+PREMIUM.*?(\d+)\s+members?.*?AED\s+([\d,]+)', text, re.I | re.DOTALL)
    if m:
        totals['members']       = int(m.group(1))
        totals['total_premium'] = int(m.group(2).replace(',', ''))
    m = re.search(r'BASMAH.*?AED\s+([\d,]+)', text, re.I | re.DOTALL)
    if m:
        totals['basmah'] = int(m.group(1).replace(',', ''))
    m = re.search(r'VALUE\s+ADDED\s+TAX.*?AED\s+([\d,]+)', text, re.I | re.DOTALL)
    if m:
        totals['vat'] = int(m.group(1).replace(',', ''))
    all_aed = re.findall(r'AED\s+([\d,]+)', text)
    if all_aed:
        amounts = sorted([int(a.replace(',', '')) for a in all_aed], reverse=True)
        totals['grand_total'] = amounts[0]
    return totals

# ── PDF Rates Parser (Claude Vision) ─────────────────────────────────────────
def _find_rate_table_pages(pdf_bytes, max_pages=4):
    """
    Return indices of pages that contain the rate table (age brackets, premiums).
    Falls back to the last few pages if none found via keywords.
    """
    KEYWORDS = ['age range', 'premium', 'male', 'female', 'subtotal',
                'category', 'maternity', 'average', 'total premium']
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    page_count = len(doc)
    scores = []
    for i in range(page_count):
        text = doc[i].get_text().lower()
        score = sum(1 for kw in KEYWORDS if kw in text)
        scores.append((score, i))
    doc.close()

    # Sort by score descending, take top pages, return sorted by page order
    scores.sort(key=lambda x: -x[0])
    top = sorted([idx for _, idx in scores[:max_pages] if scores[0][0] > 2])
    if not top:
        # Fallback: last few pages (rate tables are usually near the end)
        top = list(range(max(0, page_count - max_pages), page_count))
    return top


def parse_rates_pdf(pdf_bytes, plan=''):
    """Use Claude vision to extract age bracket rates, quote totals, and member count from a PDF rates table."""
    client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY', ''))

    rate_pages = _find_rate_table_pages(pdf_bytes, max_pages=4)

    content = []
    for i in rate_pages:
        img_b64, _ = pdf_page_image(pdf_bytes, i, scale=1.5)
        content.append({
            'type': 'image',
            'source': {'type': 'base64', 'media_type': 'image/png', 'data': img_b64},
        })

    if plan.lower() == 'healthxclusive':
        maternity_rule = (
            '- maternity_rate: look for "Additional Maternity Premium" or\n'
            '  "Additional Maternity Premium / Married Females". The FIRST numeric value\n'
            '  shown for that label is the PER-ELIGIBLE-FEMALE rate (e.g. AED 5,264).\n'
            '  A "Total Maternity Premium" line (e.g. AED 21,055) is the SUM — do NOT use it.\n'
            '  Use 0 only if the label is entirely absent.\n'
        )
    else:
        maternity_rule = (
            '- maternity_rate: each category table ends with a summary block:\n'
            '    "Maternity Premium  AED X"  ← total (may be 0)\n'
            '    "Census  N"                 ← count of eligible females\n'
            '    "Average  AED Y"            ← per-capita loading ← USE THIS VALUE\n'
            '  Extract Y as maternity_rate. NOT X. Use 0 only if Average row is absent.\n'
            '  Typical range 3,000–8,000 AED.\n'
        )

    content.append({'type': 'text', 'text': (
        'Extract the following from this insurance rates/quote PDF and return ONLY valid JSON '
        '(no markdown fences, no explanation):\n\n'
        '{\n'
        '  "company_name": "string or empty",\n'
        '  "start_date": "YYYY-MM-DD or empty",\n'
        '  "confirmed_quote": <net premium before VAT as number, 0 if not found>,\n'
        '  "members": <total insured count as integer, 0 if not found>,\n'
        '  "rate_columns": <"2col" if only Male/Female, "4col" if Employee+Dependent split>,\n'
        '  "categories": {\n'
        '    "A": {\n'
        '      "brackets": [\n'
        '        {"label": "18-25", "age_lo": 18, "age_hi": 25,\n'
        '         "male": 7665.00, "female": 8410.00,\n'
        '         "dep_male": 8332.00, "dep_female": 8783.00}\n'
        '      ],\n'
        '      "maternity_rate": 4500.00\n'
        '    }\n'
        '  }\n'
        '}\n\n'
        '════ CRITICAL: COLUMN ORDER ════\n'
        'The rate table columns from LEFT TO RIGHT are ALWAYS in this fixed order:\n'
        '  Column 1: Employee Male PREMIUM  ← extract as "male"\n'
        '  Column 2: Employee Male COUNT    ← IGNORE (small integer, e.g. 0,1,2,3)\n'
        '  Column 3: Employee Female PREMIUM ← extract as "female"\n'
        '  Column 4: Employee Female COUNT  ← IGNORE\n'
        '  Column 5: Dependent Male PREMIUM ← extract as "dep_male"\n'
        '  Column 6: Dependent Male COUNT   ← IGNORE\n'
        '  Column 7: Dependent Female PREMIUM ← extract as "dep_female"\n'
        '  Column 8: Dependent Female COUNT ← IGNORE\n\n'
        'PREMIUM values are large numbers (typically 1,000–100,000 AED for UAE insurance).\n'
        'COUNT values are small integers (0, 1, 2, 3 … representing number of members).\n'
        'NEVER use a COUNT column value as a premium rate.\n'
        'If you are unsure, pick the LARGER of the two adjacent values — the large one is always the premium.\n\n'
        'AGE BANDS WITH NO EMPLOYEE:\n'
        '  For age bands 0-10 and 11-17, employee (principal) premiums are often 0 or blank.\n'
        '  Set male=0, female=0 for those bands.\n'
        '  The DEPENDENT columns (dep_male, dep_female) may still have large non-zero values — always read them.\n\n'
        '════ OTHER RULES ════\n'
        '- Extract ALL age bracket rows — do not skip any\n'
        '- If only 2 rate columns (no Employee/Dependent split): male/female = rates; dep_male=male, dep_female=female; rate_columns="2col"\n'
        '- Extract all categories (A, B, C …) if multiple exist\n'
        '- confirmed_quote = net premium BEFORE VAT (not grand total)\n'
        '- members = total insured member count\n'
        + maternity_rule +
        '- age_lo and age_hi must be integers; use 99 for the upper bound of the last bracket\n'
        '- Return ONLY the JSON object'
    )})

    response = client.messages.create(
        model='claude-haiku-4-5-20251001',
        max_tokens=4096,
        messages=[{'role': 'user', 'content': content}],
    )

    raw = response.content[0].text.strip()
    raw = re.sub(r'^```(?:json)?\s*', '', raw)
    raw = re.sub(r'\s*```$', '', raw)
    # Fix common JSON issues from LLM output
    raw = re.sub(r',\s*([}\]])', r'\1', raw)   # trailing commas
    raw = re.sub(r'//[^\n]*', '', raw)           # single-line comments
    result = json.loads(raw)

    result.setdefault('company_name', '')
    result.setdefault('start_date', '')
    result['confirmed_quote'] = float(result.get('confirmed_quote') or 0)
    result['members']         = int(result.get('members') or 0)
    result.setdefault('categories', {})

    four_col = result.get('rate_columns') == '4col'

    # ── Sanity-check extracted rates ──────────────────────────────────────────
    # UAE health insurance premiums are always ≥ 500 AED.
    # Any non-zero value below this threshold is almost certainly a census count,
    # not a premium — clamp it to 0 to prevent count columns polluting rates.
    RATE_MIN = 500.0

    # Also compute per-category median of all non-zero rates so we can catch
    # outliers that are tiny relative to the rest (e.g., "1" slipping through).
    for cat_data in result['categories'].values():
        cat_data.setdefault('maternity_rate', 0.0)
        cat_data.setdefault('brackets', [])

        # Collect all non-zero candidate rates for this category
        all_vals = []
        for b in cat_data['brackets']:
            for key in ('male', 'female', 'dep_male', 'dep_female'):
                v = float(b.get(key, 0) or 0)
                if v > 0:
                    all_vals.append(v)

        # Dynamic threshold: anything < 1% of the median is a mis-read count
        if all_vals:
            all_vals.sort()
            median_rate = all_vals[len(all_vals) // 2]
            dynamic_min = max(RATE_MIN, median_rate * 0.01)
        else:
            dynamic_min = RATE_MIN

        def _clamp_rate(v):
            v = float(v or 0)
            return 0.0 if 0 < v < dynamic_min else v

        for b in cat_data['brackets']:
            b['male']   = _clamp_rate(b.get('male', 0))
            b['female'] = _clamp_rate(b.get('female', 0))
            if four_col:
                b['dep_male']  = _clamp_rate(b.get('dep_male', 0))
                b['dep_female']= _clamp_rate(b.get('dep_female', 0))
            else:
                b['dep_male']  = float(b.get('dep_male') or b['male'])
                b['dep_female']= float(b.get('dep_female') or b['female'])
            b['age_lo'] = int(b.get('age_lo', 0))
            b['age_hi'] = int(b.get('age_hi', 99))

    result['four_col_rates'] = four_col

    # Fields expected by the frontend tool_data shape
    result['product']    = plan or ''
    result['underwriter'] = ''
    result['fees']       = {}
    result['flat_fees']  = {}

    return result


# ── HealthXclusive Tool Parser ────────────────────────────────────────────────
def _safe_num(v, default=0.0):
    """Return float from a cell value, or default if None / error string."""
    try:
        if v is None:
            return default
        if isinstance(v, str) and ('#' in v or not v.strip() or v.strip().upper() == 'N/A'):
            return default
        return float(v)
    except Exception:
        return default

def maybe_decrypt(file_bytes, password=''):
    """Decrypt an Office file if encrypted; return original bytes if not."""
    try:
        import msoffcrypto
        office_file = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
        if office_file.is_encrypted():
            decrypted = io.BytesIO()
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            return decrypted.getvalue()
    except Exception:
        pass
    return file_bytes


def parse_healthxclusive_tool(excel_bytes):
    """
    Parse the HealthXclusive Tool Excel's 'Premium Summary' sheet.
    Returns a structured dict with policy info, fees and per-category rate brackets.
    """
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        xl = openpyxl.load_workbook(io.BytesIO(maybe_decrypt(excel_bytes)), data_only=True)

    if 'Premium Summary' not in xl.sheetnames:
        raise ValueError("'Premium Summary' sheet not found. Please upload the HealthXclusive Tool Excel.")

    ws = xl['Premium Summary']
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]

    result = {
        'company_name': '', 'product': '', 'start_date': '', 'end_date': '',
        'underwriter': '', 'validated_by': '', 'policy_type': 'New Business',
        'payment_frequency': 'Annual', 'endorsement_frequency': 'Quarterly',
        'confirmed_quote': 0.0, 'members': 0,
        'flat_fees': {'basmah': 19, 'hcv': 18, 'trudoc': 12, 'slash': 0},
        'fees': {
            'broker':  {'hsb': 0.0, 'lsb': 0.0},
            'nas':     {'hsb': 0.0, 'lsb': 0.0},
            'qic':     {'hsb': 0.0, 'lsb': 0.0},
            'healthx': {'hsb': 0.0, 'lsb': 0.0},
            'levy':    {'hsb': 0.0, 'lsb': 0.0},
            'total':   {'hsb': 0.0, 'lsb': 0.0},
        },
        'categories': {}
    }

    for row in rows:
        if not any(v is not None for v in row):
            continue

        def g(i):
            return row[i] if i < len(row) else None

        r0  = str(g(0) or '').strip()
        r0l = r0.lower()

        # ── Policy info (col A = label, col B = value) ──
        if r0l == 'product':
            result['product'] = str(g(1) or '').strip()
        elif 'policy holder' in r0l:
            result['company_name'] = str(g(1) or '').strip()
        elif 'policy start date' in r0l:
            v = g(1)
            result['start_date'] = v.strftime('%Y-%m-%d') if isinstance(v, datetime) else str(v or '')
        elif 'policy end date' in r0l:
            v = g(1)
            result['end_date'] = v.strftime('%Y-%m-%d') if isinstance(v, datetime) else str(v or '')
        elif 'policy type' in r0l:
            result['policy_type'] = str(g(1) or '').strip()
        elif 'inception premium' in r0l:
            result['payment_frequency'] = str(g(1) or '').strip()
        elif 'endorsement frequency' in r0l:
            result['endorsement_frequency'] = str(g(1) or '').strip()
        elif r0l == 'underwriter':
            result['underwriter'] = str(g(1) or '').strip()
        elif 'validated by' in r0l:
            result['validated_by'] = str(g(1) or '').strip()

        # ── Fees section (label in col N=13, HSB in col O=14, LSB in col Q=16) ──
        r13 = str(g(13) or '').strip().lower()

        if 'dha basmah' in r13:
            result['flat_fees']['basmah'] = int(_safe_num(g(14), 19))
        elif 'dha hcv' in r13:
            result['flat_fees']['hcv'] = int(_safe_num(g(14), 18))
        elif 'trudoc' in r13:
            result['flat_fees']['trudoc'] = int(_safe_num(g(14), 12))
        elif 'slash' in r13:
            result['flat_fees']['slash'] = int(_safe_num(g(14), 0))
        elif r13 == 'broker':
            result['fees']['broker'] = {'hsb': _safe_num(g(14)), 'lsb': _safe_num(g(16))}
            cq = g(20)
            if cq is not None and _safe_num(cq) > 0:
                result['confirmed_quote'] = _safe_num(cq)
        elif 'nas' in r13 or 'tpa' in r13:
            result['fees']['nas'] = {'hsb': _safe_num(g(14)), 'lsb': _safe_num(g(16))}
            mb = g(20)
            if mb is not None and _safe_num(mb) > 0:
                result['members'] = int(_safe_num(mb))
        elif 'qic' in r13:
            result['fees']['qic'] = {'hsb': _safe_num(g(14)), 'lsb': _safe_num(g(16))}
        elif 'healthx' in r13 or 'health x' in r13:
            result['fees']['healthx'] = {'hsb': _safe_num(g(14)), 'lsb': _safe_num(g(16))}
        elif 'insurance levy' in r13:
            result['fees']['levy'] = {'hsb': _safe_num(g(14)), 'lsb': _safe_num(g(16))}
        elif 'total fees' in r13:
            result['fees']['total'] = {'hsb': _safe_num(g(14)), 'lsb': _safe_num(g(16))}

        # ── Rate rows: 'Cat A (EMPDEP)', 'Cat B (EMPDEP)', maternity rows ──
        if r0.startswith('Cat ') and len(r0) >= 5:
            cat_letter = r0[4].upper()
            is_mat = 'MAT' in r0.upper() and 'EMPDEPMAT' in r0.upper() or \
                     (str(g(15) or '').lower().strip() == 'additional maternity premium')
            is_dep = 'EMPDEP' in r0 and not is_mat

            if cat_letter not in result['categories']:
                result['categories'][cat_letter] = {'brackets': [], 'maternity_rate': 0}

            age_band     = str(g(1) or '').strip()
            male_gross   = g(6)
            female_gross = g(7)

            if is_dep and age_band:
                m_ok = isinstance(male_gross, (int, float))
                f_ok = isinstance(female_gross, (int, float))
                if m_ok or f_ok:
                    parts = age_band.split('-')
                    try:
                        lo = int(parts[0])
                        hi = int(parts[1]) if len(parts) > 1 else 99
                    except Exception:
                        continue
                    result['categories'][cat_letter]['brackets'].append({
                        'label':  age_band,
                        'age_lo': lo,
                        'age_hi': hi,
                        'male':   round(float(male_gross), 2) if m_ok else 0.0,
                        'female': round(float(female_gross), 2) if f_ok else 0.0,
                    })
            elif is_mat:
                if isinstance(female_gross, (int, float)):
                    result['categories'][cat_letter]['maternity_rate'] = round(float(female_gross), 2)

    # Remove categories with no valid (non-zero) rate brackets
    result['categories'] = {
        k: v for k, v in result['categories'].items()
        if v['brackets'] and any(b['male'] > 0 or b['female'] > 0 for b in v['brackets'])
    }

    return result


def parse_openx_tool(tool_bytes):
    """Parse OpenX quote tool Excel file (Premium Summary sheet).

    Structure differs from Healthxclusive:
    - Category headers: 'CAT A', 'CAT B', 'CATEGORY C', … in column A
    - Column headers row has 'Age Range' in col A, 'Premium' in col B (male) and col D (female)
    - Bracket rows: col A = age range label, col B = male rate, col D = female rate
    - Employees and dependents share the same rate — extracted once, applied to both
    """
    wb = openpyxl.load_workbook(io.BytesIO(maybe_decrypt(tool_bytes)), data_only=True)
    ws = wb['Premium Summary']

    result = {
        'company_name':   '',
        'product':        'OpenX',
        'start_date':     '',
        'end_date':       '',
        'confirmed_quote': 0.0,
        'members':        0,
        'flat_fees':      {'basmah': 37, 'hcv': 0, 'trudoc': 0, 'slash': 0},
        'fees':           {},
        'categories':     {},
    }

    # Extract company name and start date from rows 1-3
    for r in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        label = str(r[0] or '').strip().rstrip(':')
        if label == 'Policyholder':
            result['company_name'] = str(r[2] or '').strip()
        elif label == 'Start Date':
            v = r[2]
            if hasattr(v, 'strftime'):
                result['start_date'] = v.strftime('%Y-%m-%d')

    # I8 = confirmed quote premium, I13 = grand total incl. VAT (col I = index 8)
    v_i8 = ws.cell(row=8,  column=9).value
    v_i13 = ws.cell(row=13, column=9).value
    if isinstance(v_i8,  (int, float)): result['confirmed_quote'] = round(float(v_i8),  2)
    if isinstance(v_i13, (int, float)): result['grand_total']     = round(float(v_i13), 2)

    current_cat  = None
    in_brackets  = False
    total_male   = 0
    total_female = 0

    for row in ws.iter_rows(values_only=True):
        a = str(row[0] or '').strip()

        # Detect category header: 'CAT A', 'CAT B', 'CATEGORY C', etc.
        cat_match = re.match(r'^CAT(?:EGORY)?\s+([A-Z])', a, re.IGNORECASE)
        if cat_match:
            current_cat = cat_match.group(1).upper()
            in_brackets = False
            result['categories'][current_cat] = {'brackets': [], 'maternity_rate': 0}
            continue

        # Column headers row signals start of bracket data
        if a == 'Age Range':
            in_brackets = True
            continue

        # End of bracket section — accumulate member counts
        if a == 'Subtotals':
            in_brackets = False
            m_count = row[2]   # Column C — male subtotal
            f_count = row[4]   # Column E — female subtotal
            if isinstance(m_count, (int, float)): total_male   += int(m_count)
            if isinstance(f_count, (int, float)): total_female += int(f_count)
            continue

        # Parse bracket rows
        if in_brackets and current_cat and a:
            male_rate   = row[1]   # Column B — male premium
            female_rate = row[3]   # Column D — female premium

            m_ok = isinstance(male_rate,   (int, float))
            f_ok = isinstance(female_rate, (int, float))
            if not (m_ok or f_ok):
                continue  # skip #DIV/0! or empty rows

            parts = a.split('-')
            try:
                lo = int(parts[0])
                hi = int(parts[1]) if len(parts) > 1 else 99
            except (ValueError, IndexError):
                continue

            result['categories'][current_cat]['brackets'].append({
                'label':  a,
                'age_lo': lo,
                'age_hi': hi,
                'male':   round(float(male_rate),   2) if m_ok else round(float(female_rate), 2),
                'female': round(float(female_rate), 2) if f_ok else round(float(male_rate),   2),
            })

    # Remove categories with no valid brackets
    result['categories'] = {
        k: v for k, v in result['categories'].items()
        if v['brackets']
    }

    result['members'] = total_male + total_female

    return result


# ── Census Parsing ────────────────────────────────────────────────────────────
def detect_header_row(rows):
    search = {'dob', 'date of birth', 'gender', 'relation', 'category', 'marital'}
    for i, row in enumerate(rows):
        text = ' '.join(str(v).lower() for v in row if v)
        if sum(1 for t in search if t in text) >= 2:
            return i
    return 0

def _is_notes_row(rows, header_idx, dob_col):
    """Return True if the row after the header is a notes/example row, not real data.
    Some census templates include a sample row (e.g. 'DD/MM/YYYY', 'Male/Female')
    immediately below the header. Others (like Yamaha) jump straight into data.
    We check the DOB cell: if it can't be parsed as a real date it's a notes row."""
    next_idx = header_idx + 1
    if next_idx >= len(rows):
        return False
    row = rows[next_idx]
    if dob_col >= len(row):
        return True
    val = row[dob_col]
    if val is None:
        return True
    s = str(val).strip().lower()
    if not s:
        return True
    placeholders = ('dd/', 'mm/', 'yyyy', 'date of birth', 'date', 'example', '(')
    if any(p in s for p in placeholders):
        return True
    try:
        pd.to_datetime(str(val))
        return False   # parseable real date → this is real data, not a notes row
    except Exception:
        return True

def detect_col_map(headers):
    h = [str(v).lower().strip() if v else '' for v in headers]
    col_map = {}

    for i, c in enumerate(h):
        if 'full name' in c or 'full_name' in c:
            col_map['full_name'] = i; break
    if 'full_name' not in col_map:
        for i, c in enumerate(h):
            if 'first' in c and 'name' in c:
                col_map['first_name'] = i; break
        for i, c in enumerate(h):
            if ('last' in c or 'family' in c or 'sur' in c) and 'name' in c:
                col_map['last_name'] = i; break
        if 'first_name' not in col_map:
            for i, c in enumerate(h):
                if 'name' in c and i not in col_map.values():
                    col_map['full_name'] = i; break

    for i, c in enumerate(h):
        if 'dob' in c or 'birth' in c:
            col_map['dob'] = i; break
    for i, c in enumerate(h):
        if 'gender' in c or c == 'sex':
            col_map['gender'] = i; break
    for i, c in enumerate(h):
        if 'marital' in c:
            col_map['marital'] = i; break
    for i, c in enumerate(h):
        if 'relation' in c or 'dependency' in c or 'membertype' in c or c in ('dep', 'type'):
            col_map['relation'] = i; break
    for i, c in enumerate(h):
        if c == 'category' or (c.startswith('cat') and len(c) <= 8):
            col_map['category'] = i; break
    for i, c in enumerate(h):
        if 'visa' in c and 'issuance' in c:
            col_map['emirate'] = i; break
        elif 'emirate' in c and ('visa' in c or 'issu' in c):
            col_map['emirate'] = i; break
    return col_map

def calculate_alb(dob, start_date):
    if isinstance(dob, datetime):
        dob = dob.date()
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    elif isinstance(start_date, datetime):
        start_date = start_date.date()
    age = start_date.year - dob.year
    if (start_date.month, start_date.day) < (dob.month, dob.day):
        age -= 1
    return max(0, age)

def calculate_anb(dob, start_date):
    """Age Next Birthday = Age Last Birthday + 1 (used for Healthx only)."""
    return calculate_alb(dob, start_date) + 1

def parse_census(file_bytes, filename, start_date_str, age_method='alb'):
    if filename.lower().endswith('.csv'):
        df = pd.read_csv(io.BytesIO(file_bytes))
        all_rows  = [list(df.columns)] + df.values.tolist()
        header_idx = 0
        data_start = 1
    elif filename.lower().endswith('.xls'):
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None, engine='xlrd')
        all_rows   = df.values.tolist()
        header_idx = detect_header_row(all_rows)
        _dob_col   = detect_col_map(all_rows[header_idx]).get('dob', 999)
        data_start = header_idx + 2 if _is_notes_row(all_rows, header_idx, _dob_col) else header_idx + 1
    else:
        try:
            xl = openpyxl.load_workbook(io.BytesIO(maybe_decrypt(file_bytes)), data_only=True)
        except Exception:
            # Fallback for files saved with wrong extension or legacy format
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None, engine='xlrd')
            all_rows   = df.values.tolist()
            header_idx = detect_header_row(all_rows)
            _dob_col   = detect_col_map(all_rows[header_idx]).get('dob', 999)
            data_start = header_idx + 2 if _is_notes_row(all_rows, header_idx, _dob_col) else header_idx + 1
        else:
            # Prefer INCEP sheet but fall back to whichever sheet has data rows
            _incep = next((s for s in xl.sheetnames if 'INCEP' in s.upper()), None)
            _candidates = ([_incep] if _incep else []) + [s for s in xl.sheetnames if s != _incep]
            all_rows, header_idx, _dob_col, data_start = None, 0, 999, 1
            for _sname in _candidates:
                _ws = xl[_sname]
                _rows = [[cell.value for cell in row] for row in _ws.iter_rows()]
                _hi   = detect_header_row(_rows)
                _dc   = detect_col_map(_rows[_hi]).get('dob', 999)
                _ds   = _hi + 2 if _is_notes_row(_rows, _hi, _dc) else _hi + 1
                if any(any(v is not None for v in r) for r in _rows[_ds:]):
                    all_rows, header_idx, _dob_col, data_start = _rows, _hi, _dc, _ds
                    break
            if all_rows is None:
                # All sheets empty — use first INCEP sheet anyway
                _ws = xl[_candidates[0]]
                all_rows   = [[cell.value for cell in row] for row in _ws.iter_rows()]
                header_idx = detect_header_row(all_rows)
                _dob_col   = detect_col_map(all_rows[header_idx]).get('dob', 999)
                data_start = header_idx + 2 if _is_notes_row(all_rows, header_idx, _dob_col) else header_idx + 1

    col_map = detect_col_map(all_rows[header_idx])
    if 'dob' not in col_map:
        raise ValueError("Cannot find Date of Birth column in census file")
    if 'emirate' not in col_map:
        raise ValueError("Cannot find 'Emirates of Visa Issuance' column in census file")

    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    members = []
    parents_excluded = 0

    for row_vals in all_rows[data_start:]:
        dob_raw = row_vals[col_map['dob']] if col_map['dob'] < len(row_vals) else None
        if dob_raw is None:
            continue
        if isinstance(dob_raw, str):
            if not dob_raw.strip() or 'passport' in dob_raw.lower() or dob_raw.startswith('('):
                continue

        try:
            if isinstance(dob_raw, (datetime, date)):
                dob = dob_raw.date() if isinstance(dob_raw, datetime) else dob_raw
            else:
                dob = pd.to_datetime(str(dob_raw)).date()
        except Exception:
            continue

        def get_val(key, default=''):
            idx = col_map.get(key)
            if idx is None or idx >= len(row_vals):
                return default
            v = row_vals[idx]
            return str(v).strip() if v is not None else default

        # Build name
        name = ''
        if 'full_name' in col_map:
            name = get_val('full_name')
            if name.startswith('=') or not name or name == 'nan':
                name = ''

        if not name and 'first_name' not in col_map and 'last_name' not in col_map:
            # QIC format fallback: columns B(1), C(2), D(3) = first, second, last name
            # Only use this when no named name columns were detected
            dob_col_idx = col_map.get('dob', -1)
            parts = []
            for idx in [1, 2, 3]:
                if idx == dob_col_idx:
                    continue  # skip DOB column
                if idx < len(row_vals) and row_vals[idx]:
                    p = str(row_vals[idx]).strip()
                    if p and p != 'nan' and not p.startswith('=') and not p.startswith('('):
                        parts.append(p)
            name = ' '.join(parts)

        if not name or name == 'nan':
            # Try first_name + last_name col_map keys
            fn = get_val('first_name')
            ln = get_val('last_name')
            parts = [p for p in [fn, ln] if p and p != 'nan']
            name = ' '.join(parts)

        if not name:
            continue

        gender        = get_val('gender', 'Unknown')
        if gender in ('nan', ''):
            gender = 'Unknown'

        marital       = get_val('marital', 'Unknown')
        if marital in ('nan', ''):
            marital = 'Unknown'

        relation      = get_val('relation', 'Employee')
        if relation in ('nan', ''):
            relation = 'Employee'

        category      = get_val('category', 'A').upper()
        if category in ('NAN', ''):
            category = 'A'

        emirate_raw = get_val('emirate', '')
        if not emirate_raw or emirate_raw.lower() in ('nan', 'none', ''):
            emirate = 'Dubai'  # default when blank or formula result not cached
        else:
            emirate = emirate_raw.strip()

        age_fn = calculate_anb if age_method == 'anb' else calculate_alb
        raw = {
            'name':           name,
            'dob':            dob.strftime('%d-%b-%Y'),
            'gender':         gender,
            'marital_status': marital,
            'relation':       relation,
            'category':       category,
            'age_alb':        age_fn(dob, start_date),
            'emirate':        emirate,
        }
        norm = normalize_member_fields(raw)
        if norm is None:
            parents_excluded += 1
            continue
        members.append(norm)

    return members, parents_excluded

# ── Census Normalisation Helpers ─────────────────────────────────────────────
_GENDER_MAP = {
    'male': 'M', 'm': 'M', 'man': 'M', 'male ': 'M',
    'female': 'F', 'f': 'F', 'woman': 'F', 'fem': 'F', 'female ': 'F',
}
_RELATION_MAP = {
    'employee': 'Employee', 'emp': 'Employee', 'principal': 'Employee',
    'insured': 'Employee', 'main': 'Employee',
    'spouse': 'Spouse', 'sp': 'Spouse', 'wife': 'Spouse', 'husband': 'Spouse', 'partner': 'Spouse',
    'child': 'Child', 'ch': 'Child', 'son': 'Child', 'daughter': 'Child',
    'dep': 'Child', 'dependent': 'Child', 'dependant': 'Child', 'kid': 'Child', 'children': 'Child',
    'parent': '_PARENT', 'father': '_PARENT', 'mother': '_PARENT',
    'par': '_PARENT', 'dad': '_PARENT', 'mom': '_PARENT', 'mum': '_PARENT',
    'father-in-law': '_PARENT', 'mother-in-law': '_PARENT',
}
_MARITAL_MAP = {
    'single': 'Single', 's': 'Single', 'unmarried': 'Single', 'bachelor': 'Single', 'bachelorette': 'Single',
    'married': 'Married', 'm': 'Married',
    'divorced': 'Divorced', 'd': 'Divorced',
    'widowed': 'Widowed', 'w': 'Widowed', 'widow': 'Widowed', 'widower': 'Widowed',
}


def normalize_member_fields(m):
    """Normalise gender, relation, marital_status, category in-place.
    Returns None if the member should be excluded (Parents not covered)."""
    gender_raw = m.get('gender', '').strip().lower()
    m['gender'] = _GENDER_MAP.get(gender_raw, 'Unknown' if not gender_raw else m.get('gender', 'Unknown'))

    relation_raw = m.get('relation', '').strip().lower()
    relation_norm = _RELATION_MAP.get(relation_raw)
    if relation_norm == '_PARENT':
        return None  # Exclude — parents are not eligible for coverage
    if relation_norm:
        m['relation'] = relation_norm
    elif not relation_raw:
        m['relation'] = 'Employee'

    marital_raw = m.get('marital_status', '').strip().lower()
    m['marital_status'] = _MARITAL_MAP.get(marital_raw, m.get('marital_status', 'Unknown') if marital_raw else 'Unknown')

    cat = m.get('category', 'A').strip().upper()
    # Handle "CAT A", "CATEGORY B", "CLASS C" prefixes
    cat = re.sub(r'^(?:CAT(?:EGORY)?|CLASS)\s*', '', cat).strip()
    m['category'] = cat if cat and cat not in ('NAN', '') else 'A'

    return m


def _parse_dob_for_sort(dob_str):
    try:
        return datetime.strptime(dob_str, '%d-%b-%Y').date()
    except Exception:
        return date(1900, 1, 1)


def sort_and_group_members(members):
    """Sort by category, then group by family (Employee → Spouse → Children desc age)."""
    by_cat = {}
    for m in members:
        cat = m.get('category', 'A')
        by_cat.setdefault(cat, []).append(m)

    result = []
    for cat in sorted(by_cat.keys()):
        # Sequential family grouping: Employee starts a new family block
        families = []
        current = None
        for m in by_cat[cat]:
            if m.get('relation') == 'Employee':
                if current is not None:
                    families.append(current)
                current = [m]
            else:
                if current is None:
                    current = []
                current.append(m)
        if current is not None:
            families.append(current)

        for family in families:
            emps    = [m for m in family if m.get('relation') == 'Employee']
            spouses = [m for m in family if m.get('relation') == 'Spouse']
            children = sorted(
                [m for m in family if m.get('relation') == 'Child'],
                key=lambda m: _parse_dob_for_sort(m.get('dob', '')),
                reverse=True,   # older children first (desc age = asc birth year?)
            )
            # desc age = largest age first = smallest dob year last → sort desc by dob = sort asc by age
            children = sorted(
                [m for m in family if m.get('relation') == 'Child'],
                key=lambda m: _parse_dob_for_sort(m.get('dob', '')),
            )  # ascending dob = oldest (largest age) first
            others  = [m for m in family if m.get('relation') not in ('Employee', 'Spouse', 'Child')]
            result.extend(emps + spouses + children + others)

    return result


def detect_duplicates(members):
    """Return list of [i, j] index pairs where members share the same normalised name and DOB."""
    seen = {}
    pairs = []
    for i, m in enumerate(members):
        key = (m.get('name', '').lower().strip(), m.get('dob', '').upper())
        if key in seen:
            pairs.append([seen[key], i])
        else:
            seen[key] = i
    return pairs


def get_census_warnings(members, duplicate_pairs):
    """Return {str(index): [warning_str, ...]} for members with issues."""
    dup_indices = {idx for pair in duplicate_pairs for idx in pair}
    warnings = {}
    for i, m in enumerate(members):
        w = []
        if m.get('gender') == 'Unknown':
            w.append('Unknown gender — check M or F')
        if m.get('marital_status') == 'Unknown':
            w.append('Unknown marital status')
        if m.get('relation') == 'Unknown':
            w.append('Unknown relation')
        if i in dup_indices:
            w.append('Possible duplicate — verify and delete if needed')
        if w:
            warnings[str(i)] = w
    return warnings


# ── Rate Lookup ───────────────────────────────────────────────────────────────
def find_bracket(age, brackets):
    for b in brackets:
        if b['age_lo'] <= age <= b['age_hi']:
            return b
    return None

def get_member_rate(member, categories_data):
    cat = member['category'].upper()
    if cat not in categories_data and categories_data:
        cat = sorted(categories_data.keys())[0]
    if cat not in categories_data:
        return 0, None, f"No rate data for category {member['category']}"
    brackets = categories_data[cat]['brackets']
    bracket  = find_bracket(member['age_alb'], brackets)
    if not bracket:
        return 0, None, f"No bracket for age {member['age_alb']}"
    is_female = member['gender'].lower().startswith('f')
    is_dependent = member.get('relation', 'Employee') in ('Spouse', 'Child')
    if is_dependent and 'dep_male' in bracket:
        rate = bracket.get('dep_female' if is_female else 'dep_male', 0)
    else:
        rate = bracket.get('female' if is_female else 'male', 0)
    return rate, bracket['label'], None

# ── Calculation Engine ────────────────────────────────────────────────────────
def calculate_premiums(members, categories_data):
    results = []
    for i, m in enumerate(members):
        rate, bracket_label, error = get_member_rate(m, categories_data)
        cat      = m['category'].upper()
        # maternity_rate is the per-capita average extracted from the PDF "Average" row
        mat_rate = float(categories_data.get(cat, {}).get('maternity_rate') or 0)
        emirate  = m.get('emirate', 'Dubai')
        mat_age_max = MAT_AGE_MAX_AUH if 'abu dhabi' in emirate.lower() else MAT_AGE_MAX_DXB
        maternity = 0.0
        if (m['gender'].lower().startswith('f')
                and m['marital_status'].lower().startswith('m')
                and MAT_AGE_MIN <= m['age_alb'] <= mat_age_max):
            maternity = mat_rate

        results.append({
            **m,
            'no':                i + 1,
            'base_premium':      rate,
            'age_bracket':       bracket_label or 'N/A',
            'maternity_premium': maternity,
            'mat_age_max':       mat_age_max,
            'basmah_fee':        BASMAH_FEE,
            'total_excl_vat':    rate + maternity,
            'error':             error or '',
        })

    net       = sum(r['base_premium']     for r in results)
    mat_total = sum(r['maternity_premium'] for r in results)
    bas_total = sum(r['basmah_fee']        for r in results)
    subtotal  = net + mat_total + bas_total
    vat       = subtotal * VAT_RATE

    return results, {
        'total_net':       net,
        'total_maternity': mat_total,
        'total_basmah':    bas_total,
        'subtotal':        subtotal,
        'vat':             vat,
        'grand_total':     subtotal + vat,
        'member_count':    len(results),
    }

# ── Excel Styles ─────────────────────────────────────────────────────────────
NAVY   = "003780"; ORANGE = "fb9b35"; SKY    = "35c5fc"
VIOLET = "8431cb"; WHITE  = "FFFFFF"; DARK   = "1a2332"
LGRAY  = "EEF2FF"; MGRAY  = "F0F4FF"; PINK   = "f1517b"
GREEN_BG = "D4EDDA"; GREEN_FG = "155724"
RED_BG   = "F8D7DA"; RED_FG   = "721c24"
ORANGE_BG = "FFF3E0"

def _f(name="Inter", bold=False, size=10, color=DARK):
    return Font(name=name, bold=bold, size=size, color=color)
def _fill(c):
    return PatternFill("solid", fgColor=c)
def _al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border(c="D1D9E6"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

# ── LSB Rate Calculator ───────────────────────────────────────────────────────
def calc_lsb_rates(form_data, has_lsb):
    broker  = float(form_data.get('rm_broker',        0) or 0)
    insurer = float(form_data.get('rm_insurer',       0) or 0)
    admin   = float(form_data.get('rm_wellx',         0) or 0)
    nas     = float(form_data.get('rm_tpa',           0) or 0)
    levy    = float(form_data.get('rm_insurance_tax', 0) or 0)

    if not has_lsb:
        return broker, insurer, admin, nas, levy, broker, insurer, admin, nas, levy

    broker_lsb  = min(broker, 5.0)
    excess      = max(0.0, broker - 5.0)
    insurer_lsb = insurer + excess / 2.0
    admin_lsb   = admin   + excess / 2.0
    return broker, insurer, admin, nas, levy, broker_lsb, insurer_lsb, admin_lsb, nas, levy


# ── Plan Themes ───────────────────────────────────────────────────────────────
THEMES = {
    'healthx':        {'primary': '1A5E20', 'light': 'E8F5E9', 'mid': '2E7D32'},
    'healthxclusive': {'primary': '7B1818', 'light': 'FDECEA', 'mid': 'B71C1C'},
    'openx':          {'primary': '003780', 'light': 'EEF2FF', 'mid': '1565C0'},
}

def get_theme(plan: str) -> dict:
    return THEMES.get((plan or '').lower().replace(' ', ''), THEMES['openx'])


# ── Mismatch Analysis ─────────────────────────────────────────────────────────
def build_mismatch_analysis(members_data, q_premium, c_premium, q_members, gross_loading_total=0.0):
    causes = []
    if q_members and abs(len(members_data) - q_members) > 0:
        causes.append(f"Member count: census {len(members_data)} vs quote {q_members}")
    error_members = [m for m in members_data if m.get('error', '')]
    if error_members:
        causes.append(f"{len(error_members)} member(s) have no matching rate bracket")
    if gross_loading_total > 0:
        causes.append(f"Gross loading AED {gross_loading_total:,.2f} included in final premium")
    if not causes:
        causes.append("Review age brackets, maternity eligibility and category assignments")
    return causes


# ── Combined Excel Generator ─────────────────────────────────────────────────
def make_combined_excel(form_data, members_data, verified_rates, maternity_rates,
                        loading_members, has_lsb, totals, quote_totals,
                        quoted_totals_calc=None, quoted_member_count=None,
                        census_diff_summary='', hide_commissions=False):
    from collections import defaultdict as _dd

    plan        = form_data.get('plan', '')
    plan_type   = form_data.get('plan_type', '')
    company     = form_data.get('company_name', '')
    broker_name = form_data.get('broker', '')
    underwriter = form_data.get('underwriter', '')
    start_date  = form_data.get('start_date', '')
    inception   = form_data.get('inception_payment') or 'Annual'
    endorse     = form_data.get('endorsement_freq') or 'Monthly'

    # Insurer / admin label based on plan
    is_openx     = plan.lower() == 'openx'
    insurer_name = 'DNI' if is_openx else 'QIC'
    admin_name   = 'Openx' if is_openx else 'Healthx'

    (broker_h, insurer_h, admin_h, nas_h, levy_h,
     broker_l, insurer_l, admin_l, nas_l, levy_l) = calc_lsb_rates(form_data, has_lsb)

    theme = get_theme(plan)
    PRI   = theme['primary']
    LGT   = theme['light']
    MID   = theme['mid']

    # Loading totals
    total_fees_frac  = (broker_h + insurer_h + admin_h + nas_h + levy_h) / 100.0
    total_gross_load = sum(float(lm.get('gross_loading', 0) or 0) for lm in loading_members)
    total_net_load   = total_gross_load * (1.0 - total_fees_frac)
    loading_vat      = total_net_load * VAT_RATE
    loading_grand    = total_net_load + loading_vat
    final_grand_total = totals['grand_total'] + loading_grand

    # Reconciliation
    is_healthx = plan.lower() == 'healthx'
    q_premium  = float(quote_totals.get('total_premium', 0) or 0)
    # Confirmed census premium (used for Sheet 2 totals)
    conf_premium = totals['total_net'] + totals['total_maternity']
    # For reconciliation: Healthx uses quoted census vs PDF quote; others use confirmed census
    if is_healthx and quoted_totals_calc is not None:
        c_premium = quoted_totals_calc['total_net'] + quoted_totals_calc['total_maternity']
    else:
        c_premium = conf_premium
    diff      = c_premium + total_net_load - q_premium
    is_match  = abs(diff) < 20.0
    q_members = int(quote_totals.get('members', 0) or 0)

    # ── Workbook ───────────────────────────────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Premium Summary'
    ws2 = wb.create_sheet('Premium per Member')

    def thin_border(c='D1D9E6'):
        s = Side(style='thin', color=c)
        return Border(left=s, right=s, top=s, bottom=s)

    def apply_borders(ws):
        tb = thin_border()
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value is not None:
                    cell.border = tb

    def pct_val(v):
        return round(v / 100.0, 6)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1: PREMIUM SUMMARY (PRD FORMAT)
    # ══════════════════════════════════════════════════════════════════════════
    ws = ws1

    def cv(r, c, val, bold=False, color=DARK, size=10, name='Inter',
           halign='left', fill_color=None, num_fmt=None, italic=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = Font(name=name, bold=bold, size=size, color=color, italic=italic)
        cell.alignment = Alignment(horizontal=halign, vertical='center')
        if fill_color:
            cell.fill = PatternFill('solid', fgColor=fill_color)
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    # Column widths for Sheet 1
    col_widths = [20, 13, 22, 22, 20, 12, 18, 3,
                  20, 14,  9, 10, 11, 16, 16, 20, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title (rows 1-2) ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    ws.merge_cells('A1:G1')
    t1 = ws.cell(row=1, column=1, value='PREMIUM SUMMARY')
    t1.font      = Font(name='Raleway', bold=True, size=16, color=WHITE)
    t1.fill      = PatternFill('solid', fgColor=PRI)
    t1.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 22
    ws.merge_cells('A2:G2')
    t2 = ws.cell(row=2, column=1, value=company.upper())
    t2.font      = Font(name='Raleway', bold=True, size=11, color=ORANGE)
    t2.fill      = PatternFill('solid', fgColor=PRI)
    t2.alignment = Alignment(horizontal='center', vertical='center')

    # ── Policy Info Block (rows 3–16) ──────────────────────────────────────────
    LABEL_FILL = LGT
    LABEL_CLR  = '374151'

    def info_label(r, c, text):
        cell = ws.cell(row=r, column=c, value=text)
        cell.font      = Font(name='Inter', bold=True, size=9.5, color=LABEL_CLR)
        cell.fill      = PatternFill('solid', fgColor=LABEL_FILL)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def info_val(r, c, text, bold=False):
        cell = ws.cell(row=r, column=c, value=text)
        cell.font      = Font(name='Inter', bold=bold, size=9.5, color=DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 3
    info_label(3, 1, 'Plan');      info_val(3, 2, plan, bold=True)
    ws.merge_cells('E3:F3')
    cv(3, 5, 'FLAT FIXED FEES', bold=True, color=PRI, halign='center', name='Raleway')

    # Row 4
    info_label(4, 1, 'Policy Holder'); info_val(4, 2, company, bold=True)
    cv(4, 5, 'DHA Basmah Fee', bold=True, color=DARK, size=9.5)
    cv(4, 6, 19, halign='center', size=9.5)
    cv(4, 7, 'Paid by client', size=9)

    # Row 5
    info_label(5, 1, 'Broker'); info_val(5, 2, broker_name)
    cv(5, 5, 'DHA HCV Fee', bold=True, color=DARK, size=9.5)
    cv(5, 6, 18, halign='center', size=9.5)
    cv(5, 7, 'Paid by client', size=9)

    # Row 6 — Start Date (DD MMM YYYY)
    info_label(6, 1, 'Policy Start Date')
    try:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
    except Exception:
        start_dt = start_date
    start_cell = ws.cell(row=6, column=2, value=start_dt)
    start_cell.font          = Font(name='Inter', size=9.5, color=DARK)
    start_cell.alignment     = Alignment(horizontal='center', vertical='center')
    start_cell.number_format = 'DD MMM YYYY'
    cv(6, 5, 'TruDoc Fee', bold=True, color=DARK, size=9.5)
    cv(6, 6, 12, halign='center', size=9.5)
    cv(6, 7, 'Paid by Healthx', size=9)

    # Row 7 — End Date formula (DD MMM YYYY)
    info_label(7, 1, 'Policy End Date')
    end_cell = ws.cell(row=7, column=2)
    end_cell.value          = '=DATE(YEAR(B6)+1,MONTH(B6),DAY(B6))-1'
    end_cell.font           = Font(name='Inter', size=9.5, color=DARK)
    end_cell.alignment      = Alignment(horizontal='center', vertical='center')
    end_cell.number_format  = 'DD MMM YYYY'
    cv(7, 5, 'Slash Data Fee', bold=True, color=DARK, size=9.5)
    cv(7, 6, 25, halign='center', size=9.5)

    # Row 8
    info_label(8, 1, 'Policy Type'); info_val(8, 2, plan_type)

    # Row 9
    cv(9, 6, 'FEES IN %', bold=True, color=PRI, halign='center', name='Raleway', size=9.5)

    # Row 10
    info_label(10, 1, 'Payment Mode'); info_val(10, 2, inception)
    cv(10, 6, 'HSB', bold=True, color=PRI, halign='center', name='Raleway', size=9.5)
    if has_lsb:
        cv(10, 7, 'LSB', bold=True, color=PRI, halign='center', name='Raleway', size=9.5)

    # Row 11
    info_label(11, 1, 'Endorsement Payment Mode'); info_val(11, 2, endorse)
    cv(11, 5, 'Broker', bold=True, color=DARK, size=9.5)
    cv(11, 6, pct_val(broker_h), halign='center', size=9.5, num_fmt='0.00%')
    if has_lsb:
        cv(11, 7, pct_val(broker_l), halign='center', size=9.5, num_fmt='0.00%')

    # Row 12
    cv(12, 5, f'{insurer_name} (Insurer)', bold=True, color=DARK, size=9.5)
    cv(12, 6, pct_val(insurer_h), halign='center', size=9.5, num_fmt='0.00%')
    if has_lsb:
        cv(12, 7, pct_val(insurer_l), halign='center', size=9.5, num_fmt='0.00%')

    # Row 13
    info_label(13, 1, 'Prepared by'); info_val(13, 2, underwriter)
    cv(13, 5, f'{admin_name} (Admin)', bold=True, color=DARK, size=9.5)
    cv(13, 6, pct_val(admin_h), halign='center', size=9.5, num_fmt='0.00%')
    if has_lsb:
        cv(13, 7, pct_val(admin_l), halign='center', size=9.5, num_fmt='0.00%')

    # Row 14
    cv(14, 5, 'NAS (TPA)', bold=True, color=DARK, size=9.5)
    cv(14, 6, pct_val(nas_h), halign='center', size=9.5, num_fmt='0.00%')
    if has_lsb:
        cv(14, 7, pct_val(nas_l), halign='center', size=9.5, num_fmt='0.00%')

    # Row 15
    cv(15, 5, 'Insurance Levy', bold=True, color=DARK, size=9.5)
    cv(15, 6, pct_val(levy_h), halign='center', size=9.5, num_fmt='0.00%')
    if has_lsb:
        cv(15, 7, pct_val(levy_l), halign='center', size=9.5, num_fmt='0.00%')

    # Row 16 — TOTAL FEES with live SUM formulas
    cv(16, 5, 'TOTAL FEES', bold=True, color=PRI, name='Raleway', size=9.5, halign='center')
    sum_f = ws.cell(row=16, column=6)
    sum_f.value         = '=SUM(F11:F15)'
    sum_f.font          = Font(name='Raleway', bold=True, size=9.5, color=PRI)
    sum_f.alignment     = Alignment(horizontal='center', vertical='center')
    sum_f.number_format = '0.00%'
    if has_lsb:
        sum_g = ws.cell(row=16, column=7)
        sum_g.value         = '=SUM(G11:G15)'
        sum_g.font          = Font(name='Raleway', bold=True, size=9.5, color=PRI)
        sum_g.alignment     = Alignment(horizontal='center', vertical='center')
        sum_g.number_format = '0.00%'

    # Row heights for info block
    for r in range(3, 17):
        ws.row_dimensions[r].height = 18

    # ── Thin gray borders on info ranges ──────────────────────────────────────
    def _gray_border():
        s = Side(style='thin', color='BBBBBB')
        return Border(left=s, right=s, top=s, bottom=s)

    for rng in ('A3:B8', 'A10:B11', 'A13:B13', 'E4:F7', 'E11:E16', 'F10:G16'):
        for row_cells in ws[rng]:
            for cell in row_cells:
                cell.border = _gray_border()

    # ── Premium Table Header (row 19) ──────────────────────────────────────────
    hdr_cols = ['Category', 'Age Band', 'GROSS PREMIUM\n(MALE)', 'GROSS PREMIUM\n(FEMALE)',
                'NET PREMIUM\n(MALE)', 'NET PREMIUM\n(FEMALE)']
    for ci, h in enumerate(hdr_cols, 1):
        cell = ws.cell(row=19, column=ci, value=h)
        cell.font      = Font(name='Raleway', bold=True, size=9, color=PRI)
        cell.fill      = PatternFill('solid', fgColor=LGT)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border('AABDD6')
    ws.row_dimensions[19].height = 30

    # ── Premium Data Rows (starting row 20) ────────────────────────────────────
    DATA_CLR = '1a2332'
    EMP_FILL = LGT
    DEP_FILL = 'F5F5F5'
    MAT_FILL = 'FFF9F0'

    row = 20
    MAT_BANDS = {'A': '18 - 45'}
    DEFAULT_MAT_BAND = '18 - 50'

    for cat in sorted(verified_rates.keys()):
        brackets = verified_rates[cat]
        mat_rate = float(maternity_rates.get(cat, 0) or 0)

        # Category header
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1, value=f'Category {cat}')
        cell.font      = Font(name='Raleway', bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill('solid', fgColor=PRI)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1

        # EMP rows
        for bi, b in enumerate(brackets):
            age_lo = b.get('age_lo', 0)
            if age_lo < 18:
                continue
            age_label = f"{age_lo} - {b.get('age_hi', 99)}"
            male_g    = b.get('male', 0)
            female_g  = b.get('female', 0)
            fill_c    = EMP_FILL if bi % 2 == 0 else 'EEF2FF'

            for ci, val in enumerate([f'Cat {cat} (EMP)', age_label, male_g, female_g, None, None], 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font      = Font(name='Inter', size=9.5, color=DATA_CLR)
                cell.fill      = PatternFill('solid', fgColor=fill_c)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = thin_border()
                if ci in (3, 4):
                    cell.number_format = '#,##0.00'
            net_m = ws.cell(row=row, column=5)
            net_m.value         = f'=C{row}*(1-$F$16)'
            net_m.font          = Font(name='Inter', size=9.5, color=DATA_CLR)
            net_m.fill          = PatternFill('solid', fgColor=fill_c)
            net_m.alignment     = Alignment(horizontal='center', vertical='center')
            net_m.border        = thin_border()
            net_m.number_format = '#,##0.00'
            net_f = ws.cell(row=row, column=6)
            net_f.value         = f'=D{row}*(1-$F$16)'
            net_f.font          = Font(name='Inter', size=9.5, color=DATA_CLR)
            net_f.fill          = PatternFill('solid', fgColor=fill_c)
            net_f.alignment     = Alignment(horizontal='center', vertical='center')
            net_f.border        = thin_border()
            net_f.number_format = '#,##0.00'
            ws.row_dimensions[row].height = 16
            row += 1

        # DEP rows (all brackets)
        for bi, b in enumerate(brackets):
            age_lo    = b.get('age_lo', 0)
            age_label = f"{age_lo} - {b.get('age_hi', 99)}"
            # Use dep_male/dep_female when available (4-col tables), else fall back to principal rates
            male_g    = b.get('dep_male') or b.get('male', 0)
            female_g  = b.get('dep_female') or b.get('female', 0)
            fill_c    = DEP_FILL if bi % 2 == 0 else 'EBEBEB'

            for ci, val in enumerate([f'Cat {cat} (DEP)', age_label, male_g, female_g, None, None], 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font      = Font(name='Inter', size=9.5, color=DATA_CLR)
                cell.fill      = PatternFill('solid', fgColor=fill_c)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = thin_border()
                if ci in (3, 4):
                    cell.number_format = '#,##0.00'
            net_m = ws.cell(row=row, column=5)
            net_m.value         = f'=C{row}*(1-$F$16)'
            net_m.font          = Font(name='Inter', size=9.5, color=DATA_CLR)
            net_m.fill          = PatternFill('solid', fgColor=fill_c)
            net_m.alignment     = Alignment(horizontal='center', vertical='center')
            net_m.border        = thin_border()
            net_m.number_format = '#,##0.00'
            net_f = ws.cell(row=row, column=6)
            net_f.value         = f'=D{row}*(1-$F$16)'
            net_f.font          = Font(name='Inter', size=9.5, color=DATA_CLR)
            net_f.fill          = PatternFill('solid', fgColor=fill_c)
            net_f.alignment     = Alignment(horizontal='center', vertical='center')
            net_f.border        = thin_border()
            net_f.number_format = '#,##0.00'
            ws.row_dimensions[row].height = 16
            row += 1

        # MAT row
        if mat_rate > 0:
            mat_band = MAT_BANDS.get(cat, DEFAULT_MAT_BAND)
            ws.cell(row=row, column=1).value = f'Cat {cat} (EMPDEPMAT)'
            ws.cell(row=row, column=2).value = mat_band
            ws.cell(row=row, column=4).value = mat_rate

            for ci in range(1, 8):
                cell = ws.cell(row=row, column=ci)
                cell.font      = Font(name='Inter', size=9.5, color=DATA_CLR, italic=True)
                cell.fill      = PatternFill('solid', fgColor=MAT_FILL)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border    = thin_border()

            net_mat = ws.cell(row=row, column=6)
            net_mat.value         = f'=D{row}*(1-$F$16)'
            net_mat.font          = Font(name='Inter', size=9.5, color=DATA_CLR, italic=True)
            net_mat.fill          = PatternFill('solid', fgColor=MAT_FILL)
            net_mat.alignment     = Alignment(horizontal='center', vertical='center')
            net_mat.border        = thin_border()
            net_mat.number_format = '#,##0.00'
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=7).value = 'additional maternity premium'
            ws.cell(row=row, column=7).font  = Font(name='Inter', size=8.5, color='888888', italic=True)
            ws.cell(row=row, column=7).alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 16
            row += 1

    # ── Additional Loading Section (cols I–Q) ──────────────────────────────────
    cv(21, 9, 'Additional Loading', bold=True, color=PRI, name='Raleway', size=10)

    cv(23, 9, 'Net Loading',   bold=True, color=DARK, size=9.5)
    net_load_cell = ws.cell(row=23, column=10)
    net_load_cell.value         = '=J24*(1-F16)'
    net_load_cell.font          = Font(name='Inter', size=9.5, color=DARK)
    net_load_cell.alignment     = Alignment(horizontal='center', vertical='center')
    net_load_cell.number_format = '#,##0.00'

    cv(24, 9, 'Gross Loading', bold=True, color=DARK, size=9.5)

    # Loading member table headers at row 27
    loading_hdrs = ['Member', 'DOB', 'Gender', 'Category', 'Relation',
                    'Net Loading', 'Gross Loading', 'Declaration', 'Comments']
    for ci, h in enumerate(loading_hdrs, 9):
        cell = ws.cell(row=27, column=ci, value=h)
        cell.font      = Font(name='Raleway', bold=True, size=9, color=WHITE)
        cell.fill      = PatternFill('solid', fgColor=PRI)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border(WHITE)
    ws.row_dimensions[27].height = 22

    MEMBER_ROW_START = 28
    last_loading_row = MEMBER_ROW_START - 1

    for li, lm in enumerate(loading_members):
        r      = MEMBER_ROW_START + li
        fill_c = LGT if li % 2 == 0 else 'FFFFFF'
        dob_str = str(lm.get('dob', ''))

        vals = [lm.get('name', ''), dob_str, lm.get('gender', ''),
                lm.get('category', ''), lm.get('relation', ''),
                None,
                lm.get('gross_loading', 0),
                lm.get('diagnosis', ''),
                lm.get('notes', '')]

        for ci, val in enumerate(vals, 9):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font      = Font(name='Inter', size=9.5, color=DATA_CLR)
            cell.fill      = PatternFill('solid', fgColor=fill_c)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = thin_border()
            if ci == 15:
                cell.number_format = '#,##0.00'

        # Net Loading formula: =O{row}*(1-$F$16)
        net_cell = ws.cell(row=r, column=14)
        net_cell.value         = f'=O{r}*(1-$F$16)'
        net_cell.font          = Font(name='Inter', size=9.5, color=DATA_CLR)
        net_cell.fill          = PatternFill('solid', fgColor=fill_c)
        net_cell.alignment     = Alignment(horizontal='center', vertical='center')
        net_cell.border        = thin_border()
        net_cell.number_format = '#,##0.00'

        ws.row_dimensions[r].height = 18
        last_loading_row = r

    if loading_members:
        total_r = last_loading_row + 1
        for ci in range(9, 18):
            cell = ws.cell(row=total_r, column=ci)
            cell.font   = Font(name='Raleway', bold=True, size=9.5, color=PRI)
            cell.border = thin_border()
        ws.cell(row=total_r, column=14).value         = f'=SUM(N{MEMBER_ROW_START}:N{last_loading_row})'
        ws.cell(row=total_r, column=14).number_format = '#,##0.00'
        ws.cell(row=total_r, column=15).value         = f'=SUM(O{MEMBER_ROW_START}:O{last_loading_row})'
        ws.cell(row=total_r, column=15).number_format = '#,##0.00'

        j24 = ws.cell(row=24, column=10)
        j24.value         = f'=O{total_r}'
        j24.font          = Font(name='Inter', size=9.5, color=DARK)
        j24.alignment     = Alignment(horizontal='center', vertical='center')
        j24.number_format = '#,##0.00'
    else:
        j24 = ws.cell(row=24, column=10)
        j24.value         = 0
        j24.font          = Font(name='Inter', size=9.5, color='888888', italic=True)
        j24.alignment     = Alignment(horizontal='center', vertical='center')
        j24.number_format = '#,##0.00'

    # Footer
    footer_row = row + 2
    ws.cell(row=footer_row, column=1).value = f'Created: {datetime.now().strftime("%d %b %Y  %H:%M")}'
    ws.cell(row=footer_row, column=1).font  = Font(name='Inter', size=8, color='9aa5b4', italic=True)

    apply_borders(ws1)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: PREMIUM PER MEMBER
    # ══════════════════════════════════════════════════════════════════════════
    ws = ws2

    # Column widths (A–Q normal / A–P OpenX — no maternity col)
    if is_openx:
        mem_widths = [5, 34, 14, 9, 9, 14, 14, 14, 9, 14, 14, 14, 22, 22, 22, 22]  # 16 cols
    else:
        mem_widths = [5, 34, 14, 9, 9, 14, 14, 14, 9, 14, 14, 14, 14, 22, 22, 22, 22]  # 17 cols
    for i, w in enumerate(mem_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_col_letter = 'P' if is_openx else 'Q'

    # Title
    ws.row_dimensions[1].height = 42
    ws.merge_cells(f'A1:{last_col_letter}1')
    t1m = ws.cell(row=1, column=1, value='PREMIUM PER MEMBER')
    t1m.font      = Font(name='Raleway', bold=True, size=16, color=WHITE)
    t1m.fill      = PatternFill('solid', fgColor=PRI)
    t1m.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 22
    ws.merge_cells(f'A2:{last_col_letter}2')
    t2m = ws.cell(row=2, column=1, value=company.upper())
    t2m.font      = Font(name='Raleway', bold=True, size=11, color=ORANGE)
    t2m.fill      = PatternFill('solid', fgColor=PRI)
    t2m.alignment = Alignment(horizontal='center', vertical='center')

    # Summary bar (row 3) — value cells filled with formulas after member rows exist (auditing)
    mem_row = 3
    summary_row = 3
    final_premium_with_loading = totals['total_net'] + totals['total_maternity'] + total_gross_load

    if is_openx:
        summary_items = [
            ('Total Members',  False, PRI),
            ('Net Premium',    True,  PRI),
            ('Gross Loading',  True,  PRI),
            ('Final Premium',  True,  ORANGE),
        ]
    else:
        summary_items = [
            ('Total Members',      False, PRI),
            ('Net Premium',        True,  PRI),
            ('Total Maternity',    True,  MID),
            ('Gross Loading',      True,  PRI),
            ('Final Premium',      True,  ORANGE),
        ]
    max_summary_col = len(summary_items) * 2
    for i, (label, is_aed, color) in enumerate(summary_items):
        col = (i * 2) + 1
        if col + 1 <= max_summary_col:
            lc = ws.cell(row=summary_row, column=col, value=label)
            lc.font      = Font(name='Inter', bold=True, size=8, color=WHITE)
            lc.fill      = PatternFill('solid', fgColor=color)
            lc.alignment = Alignment(horizontal='center', vertical='center')
            vc = ws.cell(row=summary_row, column=col + 1, value=None)
            vc.font      = Font(name='Inter', bold=True, size=9, color=WHITE)
            vc.fill      = PatternFill('solid', fgColor=color)
            vc.alignment = Alignment(horizontal='center', vertical='center')
            if is_aed:
                vc.number_format = '"AED "#,##0'
    ws.row_dimensions[summary_row].height = 22
    mem_row += 1

    # Column headers (row 4)
    age_col_label = 'Age (ANB)' if is_healthx else 'Age (ALB)'
    if is_openx:
        hdrs2 = ['No.', 'Full Name', 'Date of Birth', 'Gender', 'Category',
                 'Relationship', 'Marital Status', 'Emirate (Visa)', age_col_label, 'Age Bracket',
                 'Premium', 'Gross Loading', 'Final Premium (incl. loading)',
                 'Notes', 'Diagnosis', 'Loading Notes']
    else:
        hdrs2 = ['No.', 'Full Name', 'Date of Birth', 'Gender', 'Category',
                 'Relationship', 'Marital Status', 'Emirate (Visa)', age_col_label, 'Age Bracket',
                 'Premium', 'Maternity Premium', 'Gross Loading', 'Final Premium (incl. loading)',
                 'Notes', 'Diagnosis', 'Loading Notes']
    for c, h in enumerate(hdrs2, 1):
        cell = ws.cell(row=mem_row, column=c, value=h)
        cell.font      = Font(name='Inter', bold=True, size=9, color=WHITE)
        cell.fill      = PatternFill('solid', fgColor=PRI)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border(WHITE)
    ws.row_dimensions[mem_row].height = 30
    mem_row += 1

    # Build loading lookup
    loading_map = {lm.get('name', '').strip().lower(): lm for lm in loading_members}

    first_data_row = mem_row

    for m in members_data:
        alt     = m['no'] % 2 == 0
        has_mat = m['maternity_premium'] > 0
        lm_data = loading_map.get(m['name'].strip().lower())
        has_load = lm_data is not None

        if has_load:
            bg = LGT
        elif has_mat:
            bg = ORANGE_BG
        elif alt:
            bg = 'F5F5F5'
        else:
            bg = WHITE

        gross_loading = float(lm_data.get('gross_loading', 0) or 0) if has_load else 0.0
        notes_parts = []
        if has_load:
            notes_parts.append('Loading applied')
        if has_mat:
            notes_parts.append('Maternity surcharge applied')
        if m.get('error', ''):
            notes_parts.append(m.get('error', ''))
        notes = ' | '.join(notes_parts)

        # Parse DOB to date object for DD MMM YYYY format
        dob_val = m['dob']
        try:
            dob_val = datetime.strptime(str(m['dob']), '%d-%b-%Y').date()
        except Exception:
            try:
                dob_val = datetime.strptime(str(m['dob']), '%Y-%m-%d').date()
            except Exception:
                pass

        r = mem_row
        base_cols = [m['no'], m['name'], dob_val, m['gender'], m['category'],
                     m['relation'], m.get('marital_status', ''), m.get('emirate', ''),
                     m['age_alb'], m['age_bracket']]
        for c, v in enumerate(base_cols, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font   = Font(name='Inter', size=9)
            cell.fill   = PatternFill('solid', fgColor=bg)
            cell.border = thin_border()
            if c == 3:
                cell.number_format = 'DD MMM YYYY'
                cell.alignment     = Alignment(horizontal='center', vertical='center')
            elif c == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        if is_openx:
            prem_cols = zip((11, 12), (m['base_premium'], gross_loading))
        else:
            prem_cols = zip((11, 12, 13), (m['base_premium'], m['maternity_premium'], gross_loading))
        for c, v in prem_cols:
            cell = ws.cell(row=r, column=c, value=v)
            cell.font          = Font(name='Inter', size=9)
            cell.fill          = PatternFill('solid', fgColor=bg)
            cell.border        = thin_border()
            cell.number_format = '#,##0.00'
            cell.alignment     = Alignment(horizontal='right', vertical='center')

        fin_col   = 13 if is_openx else 14
        fin_range = f'K{r}:L{r}' if is_openx else f'K{r}:M{r}'
        fin = ws.cell(row=r, column=fin_col, value=f'=SUM({fin_range})')
        fin.font          = Font(name='Inter', size=9)
        fin.fill          = PatternFill('solid', fgColor=bg)
        fin.border        = thin_border()
        fin.number_format = '#,##0.00'
        fin.alignment     = Alignment(horizontal='right', vertical='center')

        notes_start = 14 if is_openx else 15
        for c, v in zip((notes_start, notes_start+1, notes_start+2), (notes,
                                       lm_data.get('diagnosis', '') if has_load else None,
                                       lm_data.get('notes', '') if has_load else None)):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font   = Font(name='Inter', size=9)
            cell.fill   = PatternFill('solid', fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal='center', vertical='center')

        mem_row += 1

    last_member_row = mem_row - 1

    # Totals row (SUM formulas — audit trail)
    totals_row = mem_row
    if is_openx:
        totals_placeholder = ['', 'TOTALS', '', '', '', '', '', '', '', '', None, None, None, '', '', '']
        num_cols_set = (11, 12, 13)
    else:
        totals_placeholder = ['', 'TOTALS', '', '', '', '', '', '', '', '', None, None, None, None, '', '', '']
        num_cols_set = (11, 12, 13, 14)
    for c, v in enumerate(totals_placeholder, 1):
        cell = ws.cell(row=totals_row, column=c, value=v)
        cell.font      = Font(name='Raleway', bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill('solid', fgColor=PRI)
        cell.border    = thin_border(WHITE)
        cell.alignment = Alignment(horizontal='right' if c in num_cols_set else 'left', vertical='center')
        if c in num_cols_set:
            cell.number_format = '#,##0.00'
    if last_member_row >= first_data_row:
        ws.cell(row=totals_row, column=11, value=f'=SUM(K{first_data_row}:K{last_member_row})')
        if is_openx:
            ws.cell(row=totals_row, column=12, value=f'=SUM(L{first_data_row}:L{last_member_row})')
            ws.cell(row=totals_row, column=13, value=f'=SUM(M{first_data_row}:M{last_member_row})')
        else:
            ws.cell(row=totals_row, column=12, value=f'=SUM(L{first_data_row}:L{last_member_row})')
            ws.cell(row=totals_row, column=13, value=f'=SUM(M{first_data_row}:M{last_member_row})')
            ws.cell(row=totals_row, column=14, value=f'=SUM(N{first_data_row}:N{last_member_row})')
    else:
        for col in num_cols_set:
            ws.cell(row=totals_row, column=col, value=0)
    mem_row += 1

    # Top summary bar formulas (reference member data rows only, not the TOTALS row)
    if last_member_row >= first_data_row:
        ws.cell(row=summary_row, column=2, value=f'=COUNTA(B{first_data_row}:B{last_member_row})')
        ws.cell(row=summary_row, column=4, value=f'=SUM(K{first_data_row}:K{last_member_row})')
        if is_openx:
            # No maternity col — Gross Loading at col 6, Final Premium at col 8
            ws.cell(row=summary_row, column=6, value=f'=SUM(L{first_data_row}:L{last_member_row})')
            ws.cell(row=summary_row, column=8, value=f'=SUM(M{first_data_row}:M{last_member_row})')
        else:
            ws.cell(row=summary_row, column=6,  value=f'=SUM(L{first_data_row}:L{last_member_row})')
            ws.cell(row=summary_row, column=8,  value=f'=SUM(M{first_data_row}:M{last_member_row})')
            ws.cell(row=summary_row, column=10, value=f'=SUM(N{first_data_row}:N{last_member_row})')
    else:
        ws.cell(row=summary_row, column=2, value=0)
        ws.cell(row=summary_row, column=4, value=0)
        ws.cell(row=summary_row, column=6, value=0)
        ws.cell(row=summary_row, column=8, value=0)
        if not is_openx:
            ws.cell(row=summary_row, column=10, value=0)

    # ── Summary block (Fix 10) ─────────────────────────────────────────────────
    mem_row += 2

    total_fees_pct = broker_h + insurer_h + admin_h + nas_h + levy_h
    comm_str = (f'Broker {broker_h}% | {insurer_name} {insurer_h}% | '
                f'{admin_name} {admin_h}% | NAS {nas_h}% | Levy {levy_h}% | '
                f'Total {total_fees_pct}%')

    if is_match:
        analysis_txt = '✅ Match (within AED 20)'
    else:
        causes = build_mismatch_analysis(members_data, q_premium, c_premium, q_members, total_gross_load)
        analysis_txt = '❌ ' + ' | '.join(causes)

    if census_diff_summary:
        analysis_txt += f'\nCensus diff: {census_diff_summary}'

    summary_block = [
        ('Company Name',                  company),
        ('Broker',                        broker_name),
    ]
    if not hide_commissions:
        summary_block.append(('Commission Structure', comm_str))
    summary_block += [
        ('Confirmed Premium',             f'AED {q_premium:,.2f}'),
        ('Final Premium (incl. loading)', None),
        ('Premium Analysis',              analysis_txt),
    ]

    for label, value in summary_block:
        lbl_cell = ws.cell(row=mem_row, column=1, value=label)
        lbl_cell.font      = Font(name='Inter', bold=True, size=9.5, color='374151')
        lbl_cell.fill      = PatternFill('solid', fgColor=LGT)
        lbl_cell.alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(f'B{mem_row}:E{mem_row}')
        if label == 'Final Premium (incl. loading)':
            final_ref = 'H3' if is_openx else 'J3'
            val_cell = ws.cell(row=mem_row, column=2, value=f'="AED "&TEXT({final_ref},"#,##0.00")')
        else:
            val_cell = ws.cell(row=mem_row, column=2, value=value)
        val_cell.font      = Font(name='Inter', size=9.5, color=DARK)
        val_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        if label == 'Premium Analysis':
            fg_c = GREEN_FG if is_match else RED_FG
            bg_c = GREEN_BG if is_match else RED_BG
            lbl_cell.font = Font(name='Inter', bold=True, size=9.5, color=fg_c)
            lbl_cell.fill = PatternFill('solid', fgColor=bg_c)
            val_cell.font = Font(name='Inter', bold=True, size=9.5, color=fg_c)
            val_cell.fill = PatternFill('solid', fgColor=bg_c)

        ws.row_dimensions[mem_row].height = 18
        mem_row += 1

    apply_borders(ws2)

    # ── Wellx branding: logo in ws1 title cell + powered-by footer ────────────
    _xl_add_logo(ws1, 'A1', _WELLX_LOGO_PATH, w=110, h=36)
    footer_row2 = mem_row + 1
    ws.merge_cells(start_row=footer_row2, start_column=1,
                   end_row=footer_row2, end_column=5)
    fc = ws.cell(row=footer_row2, column=1,
                 value=f'Powered by Wellx Labs  ·  Generated {datetime.now().strftime("%d %b %Y %H:%M")}')
    fc.font      = Font(name='Inter', size=8, italic=True, color='9aa5b4')
    fc.alignment = Alignment(horizontal='center', vertical='center')
    _xl_add_logo(ws, f'F{footer_row2}', _POWERED_LOGO_PATH, w=100, h=24)

    return wb


# ── Database ──────────────────────────────────────────────────────────────────

def get_supa():
    """Return a Supabase client or None if not configured."""
    if not _DB_AVAILABLE:
        return None
    url  = os.environ.get('SUPABASE_URL', '')
    # Service role key preferred (bypasses RLS); fall back to anon key
    key  = (os.environ.get('SUPABASE_SERVICE_ROLE_KEY') or
            os.environ.get('SUPABASE_ANON_KEY') or '')
    if not url or not key:
        return None
    try:
        return create_client(url, key)
    except Exception as e:
        print(f'[DB] Supabase client error: {e}')
        return None


def _parse_dob(dob_str):
    """Parse a dob string to YYYY-MM-DD string for Supabase."""
    if not dob_str or str(dob_str) == 'None':
        return None
    for fmt in ('%d-%b-%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(str(dob_str), fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _f(v, default=0.0):
    try:
        return float(v or default)
    except (TypeError, ValueError):
        return default


def save_policy(fd, members_data, verified_rates, maternity_rates,
                loading_members, totals, quote_totals):
    """
    Persist a finalized premium summary to Supabase via REST API.
    Returns the new policy_id, or None on failure / DB not configured.
    """
    supa = get_supa()
    if not supa:
        return None
    try:
        loading_map = {lm.get('name', '').strip().lower(): lm
                       for lm in (loading_members or [])}

        q_premium   = _f(quote_totals.get('total_premium'))
        c_premium   = totals['total_net'] + totals['total_maternity']
        recon_diff  = c_premium - q_premium
        recon_match = abs(recon_diff) < 20.0

        # ── Insert policy ──────────────────────────────────────────────────
        pol_res = supa.table('policies').insert({
            'company_name':     fd.get('company_name', ''),
            'broker':           fd.get('broker', ''),
            'underwriter':      fd.get('underwriter', ''),
            'rm_person':        fd.get('rm_person', ''),
            'plan':             fd.get('plan', ''),
            'plan_type':        fd.get('plan_type', ''),
            'start_date':       fd.get('start_date') or None,
            'confirmation_date': fd.get('confirmation_date') or None,
            'recon_note':       fd.get('recon_note') or None,
            'inception_payment': fd.get('inception_payment', ''),
            'endorsement_freq': fd.get('endorsement_freq', ''),
            'has_lsb':          bool(fd.get('has_lsb', False)),
            'rm_broker':        _f(fd.get('rm_broker')),
            'rm_insurer':       _f(fd.get('rm_insurer')),
            'rm_wellx':         _f(fd.get('rm_wellx')),
            'rm_tpa':           _f(fd.get('rm_tpa')),
            'rm_insurance_tax': _f(fd.get('rm_insurance_tax')),
            'confirmed_quote':  _f(quote_totals.get('total_premium')),
            'quoted_members':   int(quote_totals.get('members', 0) or 0),
            'quote_grand_total': _f(quote_totals.get('grand_total')),
            'member_count':     totals.get('member_count', len(members_data)),
            'total_net':        totals['total_net'],
            'total_maternity':  totals['total_maternity'],
            'total_basmah':     totals['total_basmah'],
            'subtotal':         totals['subtotal'],
            'vat':              totals['vat'],
            'grand_total':      totals['grand_total'],
            'recon_match':      recon_match,
            'recon_difference': recon_diff,
        }).execute()
        policy_id = pol_res.data[0]['id']

        # ── Insert members ─────────────────────────────────────────────────
        mem_rows = []
        for m in members_data:
            lm_data    = loading_map.get(m['name'].strip().lower())
            gross_load = _f(lm_data.get('gross_loading')) if lm_data else 0.0
            final_prem = m['base_premium'] + m['maternity_premium'] + gross_load
            mem_rows.append({
                'policy_id':        policy_id,
                'member_no':        m.get('no'),
                'name':             m.get('name', ''),
                'dob':              _parse_dob(str(m.get('dob', ''))),
                'gender':           m.get('gender', ''),
                'marital_status':   m.get('marital_status', ''),
                'relation':         m.get('relation', ''),
                'category':         m.get('category', ''),
                'age_alb':          m.get('age_alb'),
                'emirate':          m.get('emirate', ''),
                'age_bracket':      m.get('age_bracket', ''),
                'base_premium':     m['base_premium'],
                'maternity_premium': m['maternity_premium'],
                'gross_loading':    gross_load,
                'final_premium':    final_prem,
            })
        # Insert in batches of 500
        for i in range(0, len(mem_rows), 500):
            supa.table('policy_members').insert(mem_rows[i:i+500]).execute()

        # ── Insert categories + brackets ───────────────────────────────────
        for cat_letter, brackets in (verified_rates or {}).items():
            mat_rate = _f((maternity_rates or {}).get(cat_letter))
            cat_res = supa.table('policy_categories').insert({
                'policy_id':     policy_id,
                'category':      cat_letter.upper(),
                'maternity_rate': mat_rate,
            }).execute()
            cat_id = cat_res.data[0]['id']

            bracket_rows = [{
                'category_id': cat_id,
                'label':       b.get('label', ''),
                'age_lo':      b.get('age_lo'),
                'age_hi':      b.get('age_hi'),
                'male_rate':   _f(b.get('male')),
                'female_rate': _f(b.get('female')),
            } for b in brackets]
            if bracket_rows:
                supa.table('policy_brackets').insert(bracket_rows).execute()

        return policy_id
    except Exception as e:
        print(f'[DB] save_policy error: {e}')
        return None


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/brand-logo')
def brand_logo():
    return send_file(BRAND_LOGO_PATH, mimetype='image/png')

@app.route('/api/upload', methods=['POST'])
def api_upload():
    plan         = request.form.get('plan', '')
    start_date   = request.form.get('start_date', '')
    census_file  = request.files.get('census')
    rates_file   = request.files.get('tool')   # PDF rates table (or xlsx fallback)
    if not census_file:
        return jsonify({'error': 'Member Census file is required'}), 400
    if not rates_file:
        return jsonify({'error': 'Rates Table PDF is required'}), 400

    census_bytes    = census_file.read()
    census_filename = census_file.filename or 'census.xlsx'

    # ── Parse rates (PDF vision for all plans) ──
    tool_data = {}
    rates_bytes = rates_file.read()
    rates_fname = (rates_file.filename or '').lower()
    try:
        if rates_fname.endswith('.pdf'):
            tool_data = parse_rates_pdf(rates_bytes, plan)
        elif rates_fname.endswith(('.xlsx', '.xls')):
            # Legacy Excel fallback
            if plan.lower() == 'openx':
                tool_data = parse_openx_tool(rates_bytes)
            else:
                tool_data = parse_healthxclusive_tool(rates_bytes)
        else:
            tool_data = parse_rates_pdf(rates_bytes, plan)
    except Exception as e:
        return jsonify({'error': f'Rates parsing error: {str(e)}'}), 400

    # Use start_date from form or from extracted tool_data
    effective_start = start_date or tool_data.get('start_date', '')
    age_method = 'anb' if plan.lower() == 'healthx' else 'alb'

    # ── Parse census ──────────────────────────────────────────────────────────
    census_members = []
    parents_excluded = 0
    if effective_start:
        try:
            census_members, parents_excluded = parse_census(
                census_bytes, census_filename, effective_start, age_method)
        except Exception as e:
            # Non-fatal: return error info in census_preview so user sees it
            census_members = []

    # Sort and group, detect duplicates, build warnings
    sorted_members = sort_and_group_members(census_members)
    dup_pairs      = detect_duplicates(sorted_members)
    warnings       = get_census_warnings(sorted_members, dup_pairs)

    # Category stats
    cat_counts = {}
    for m in sorted_members:
        cat_counts[m['category']] = cat_counts.get(m['category'], 0) + 1

    token = str(uuid.uuid4())
    _session_set(token, {
        'plan':            plan,
        'census_bytes':    census_bytes,
        'census_filename': census_filename,
        'members_preview': sorted_members,   # user-correctable
        'tool_data':       tool_data,
    })

    return jsonify({
        'token':             token,
        'tool_data':         tool_data,
        'vision_categories': tool_data.get('categories', {}),
        'four_col_rates':    tool_data.get('four_col_rates', False),
        'census_preview': {
            'members':       sorted_members,
            'warnings':      warnings,
            'duplicate_pairs': dup_pairs,
            'stats': {
                'total':            len(sorted_members),
                'categories':       cat_counts,
                'warning_count':    len(warnings),
                'parents_excluded': parents_excluded,
            },
        },
    })

@app.route('/api/update_census/<token>', methods=['POST'])
def api_update_census(token):
    """Accept user-corrected census members list and store for use in calculate."""
    stored = _session_get(token)
    if not stored:
        return jsonify({'error': 'Session not found'}), 404
    body = request.get_json(force=True) or {}
    members = body.get('members', [])
    if not isinstance(members, list):
        return jsonify({'error': 'members must be a list'}), 400
    _session_patch(token, {'members_preview': members})
    return jsonify({'ok': True, 'count': len(members)})


def compare_censuses(confirmed_members, quoted_members):
    """Compare two member lists, returning added/removed/changed dicts."""
    def _key(m):
        return (m.get('name', '').lower().strip(), m.get('category', '').upper())

    conf_map = {}
    for m in confirmed_members:
        k = _key(m)
        conf_map[k] = m

    quot_map = {}
    for m in quoted_members:
        k = _key(m)
        quot_map[k] = m

    DIFF_FIELDS = ['relation', 'dob', 'gender', 'marital_status']

    added   = [conf_map[k] for k in conf_map if k not in quot_map]
    removed = [quot_map[k] for k in quot_map if k not in conf_map]
    changed = []
    for k in conf_map:
        if k not in quot_map:
            continue
        cm = conf_map[k]
        qm = quot_map[k]
        diffs = []
        for f in DIFF_FIELDS:
            cv = cm.get(f, '')
            qv = qm.get(f, '')
            if str(cv).strip().lower() != str(qv).strip().lower():
                diffs.append({'field': f, 'confirmed': cv, 'quoted': qv})
        if diffs:
            changed.append({'name': cm.get('name'), 'category': cm.get('category'), 'changes': diffs})

    return {'added': added, 'removed': removed, 'changed': changed}


@app.route('/api/compare_census/<out_token>', methods=['POST'])
def api_compare_census(out_token):
    """Compare confirmed vs quoted census and attribute premium differences per member."""
    stored = _session_get(out_token)
    if not stored or 'members_data' not in stored:
        return jsonify({'error': 'Session not found or expired'}), 404

    census_file = request.files.get('quoted_census')
    if not census_file:
        return jsonify({'error': 'quoted_census file is required'}), 400

    start_date = stored.get('start_date') or stored.get('form_data', {}).get('start_date', '')
    age_method = stored.get('age_method', 'alb')

    try:
        quoted_members, _ = parse_census(
            census_file.read(), census_file.filename or 'quoted.xlsx',
            start_date, age_method)
    except Exception as e:
        return jsonify({'error': f'Census parsing error: {str(e)}'}), 400

    confirmed_members = stored['members_data']
    diff = compare_censuses(confirmed_members, quoted_members)

    # ── Reconstruct categories_data from stored rates ──────────────────────────
    categories_data = {}
    for cat, brackets in (stored.get('verified_rates') or {}).items():
        categories_data[cat.upper()] = {
            'brackets':      brackets,
            'maternity_rate': float((stored.get('maternity_rates') or {}).get(cat, 0) or 0),
        }

    def _member_premium(m, cats):
        """Return (base_rate, maternity, total) for a member given categories_data."""
        rate, _, _ = get_member_rate(m, cats)
        rate = float(rate or 0)
        cat  = m.get('category', '').upper()
        mat_rate = float(cats.get(cat, {}).get('maternity_rate', 0) or 0)
        emirate  = m.get('emirate', 'Dubai')
        mat_max  = MAT_AGE_MAX_AUH if 'abu dhabi' in emirate.lower() else MAT_AGE_MAX_DXB
        age      = float(m.get('age_alb') or 0)
        mat = (mat_rate
               if (m.get('gender', '').lower().startswith('f')
                   and m.get('marital_status', '').lower().startswith('m')
                   and MAT_AGE_MIN <= age <= mat_max)
               else 0.0)
        return rate, mat, rate + mat

    # ── Build quick lookups ────────────────────────────────────────────────────
    def _key(m):
        return (m.get('name', '').lower().strip(), m.get('category', '').upper())

    conf_map = {_key(m): m for m in confirmed_members}
    quot_map = {_key(m): m for m in quoted_members}

    # ── Premium impact per member ──────────────────────────────────────────────
    premium_impacts = []   # list of dicts for the frontend

    # Changed members: same person, different attributes → different rate
    for ch in diff.get('changed', []):
        name = ch.get('name', '')
        cat  = ch.get('category', '')
        cm   = conf_map.get((name.lower().strip(), cat.upper()))
        qm   = quot_map.get((name.lower().strip(), cat.upper()))
        if not cm or not qm:
            continue

        # Confirmed premium already calculated and stored
        c_base  = float(cm.get('base_premium', 0) or 0)
        c_mat   = float(cm.get('maternity_premium', 0) or 0)
        c_total = c_base + c_mat
        c_age   = cm.get('age_alb') or cm.get('age_anb') or '?'
        c_gen   = (cm.get('gender') or '?')[:1].upper()

        # Quoted premium — recalculate using stored rates
        q_base, q_mat, q_total = _member_premium(qm, categories_data)
        q_age   = qm.get('age_alb') or qm.get('age_anb') or '?'
        q_gen   = (qm.get('gender') or '?')[:1].upper()

        impact  = c_total - q_total
        if abs(impact) < 0.01 and c_age == q_age and c_gen == q_gen:
            continue  # attribute changed but same price band

        premium_impacts.append({
            'type':     'changed',
            'name':     name,
            'category': cat,
            'quoted':   {'age': q_age, 'gender': q_gen, 'premium': round(q_total, 2)},
            'confirmed':{'age': c_age, 'gender': c_gen, 'premium': round(c_total, 2)},
            'impact':   round(impact, 2),
        })

    # Added members (in confirmed, not in quoted) → they increase confirmed total
    for m in diff.get('added', []):
        c_base  = float(m.get('base_premium', 0) or 0)
        c_mat   = float(m.get('maternity_premium', 0) or 0)
        c_total = c_base + c_mat
        age     = m.get('age_alb') or m.get('age_anb') or '?'
        gen     = (m.get('gender') or '?')[:1].upper()
        premium_impacts.append({
            'type':     'added',
            'name':     m.get('name', ''),
            'category': m.get('category', ''),
            'confirmed':{'age': age, 'gender': gen, 'premium': round(c_total, 2)},
            'impact':   round(c_total, 2),
        })

    # Removed members (in quoted, not in confirmed) → they reduce confirmed vs quoted
    for m in diff.get('removed', []):
        q_base, q_mat, q_total = _member_premium(m, categories_data)
        age = m.get('age_alb') or m.get('age_anb') or '?'
        gen = (m.get('gender') or '?')[:1].upper()
        premium_impacts.append({
            'type':     'removed',
            'name':     m.get('name', ''),
            'category': m.get('category', ''),
            'quoted':   {'age': age, 'gender': gen, 'premium': round(q_total, 2)},
            'impact':   round(-q_total, 2),
        })

    # ── Overall premium reconciliation ─────────────────────────────────────────
    totals       = stored.get('totals', {})
    quote_totals = stored.get('quote_totals', {})
    calc_premium = float(totals.get('total_net', 0) or 0) + float(totals.get('total_maternity', 0) or 0)
    conf_quote   = float(quote_totals.get('total_premium', 0) or 0)
    total_gap    = calc_premium - conf_quote
    explained    = sum(p['impact'] for p in premium_impacts)
    unexplained  = total_gap - explained

    reconciliation = {
        'calc_premium':  round(calc_premium, 2),
        'conf_quote':    round(conf_quote, 2),
        'total_gap':     round(total_gap, 2),
        'explained':     round(explained, 2),
        'unexplained':   round(unexplained, 2),
    }

    diff['premium_impacts']  = premium_impacts
    diff['reconciliation']   = reconciliation

    # ── Excel diff summary (compact text) ─────────────────────────────────────
    def _fmt_m(m):
        age = m.get('age_alb') or m.get('age_anb') or '?'
        gen = (m.get('gender') or '?')[:1].upper()
        rel = m.get('relation') or '?'
        return f"{m.get('name','?')} ({age}, {gen}, {rel})"

    summary_lines = []
    if diff.get('added'):
        names = [_fmt_m(m) for m in diff['added'][:5]]
        extra = len(diff['added']) - 5
        line  = 'Added: ' + ', '.join(names) + (f' +{extra} more' if extra > 0 else '')
        summary_lines.append(line)
    if diff.get('removed'):
        names = [_fmt_m(m) for m in diff['removed'][:5]]
        extra = len(diff['removed']) - 5
        line  = 'Removed: ' + ', '.join(names) + (f' +{extra} more' if extra > 0 else '')
        summary_lines.append(line)
    if diff.get('changed'):
        ch_parts = []
        for ch in diff['changed'][:3]:
            fld = ', '.join(f"{c['field']} {c['from']}→{c['to']}" for c in ch.get('changes', []))
            ch_parts.append(f"{ch['name']} ({fld})")
        extra = len(diff['changed']) - 3
        line  = 'Changed: ' + ', '.join(ch_parts) + (f' +{extra} more' if extra > 0 else '')
        summary_lines.append(line)

    _session_patch(out_token, {'census_diff_summary': (
        ' | '.join(summary_lines) if summary_lines else 'Census identical to quoted'
    )})

    return jsonify(diff)


@app.route('/api/page/<token>/<int:page_idx>')
def api_page(token, page_idx):
    stored = _session_get(token)
    if not stored:
        return jsonify({'error': 'Session not found'}), 404
    img, total = pdf_page_image(stored['pdf_bytes'], page_idx, scale=2.0)
    return jsonify({'image': img, 'page': page_idx, 'total': total})

@app.route('/api/calculate', methods=['POST'])
def api_calculate():
    body           = request.json
    token          = body.get('token')
    form_data      = body.get('form_data', {})
    verified_rates = body.get('verified_rates', {})   # {cat: [brackets...]}
    maternity_rates= body.get('maternity_rates', {})  # {cat: amount}
    quote_totals   = body.get('quote_totals', {})

    stored = _session_get(token)
    if not stored:
        return jsonify({'error': 'Session expired. Please re-upload files.'}), 400

    start_date  = form_data.get('start_date')
    age_method  = form_data.get('age_method', 'alb')   # 'anb' for Healthx

    if not start_date:
        return jsonify({'error': 'Start date is required'}), 400

    # Use user-corrected preview if available, else re-parse census
    members = stored.get('members_preview')
    if not members:
        census_bytes    = stored.get('census_bytes', b'')
        census_filename = stored.get('census_filename', 'census.xlsx')
        try:
            members, _ = parse_census(census_bytes, census_filename, start_date, age_method)
        except Exception as e:
            return jsonify({'error': f'Census parsing error: {str(e)}'}), 400

    if not members:
        return jsonify({'error': 'No valid members found in census file'}), 400

    # Build categories_data
    categories_data = {}
    for cat, brackets in verified_rates.items():
        cat_upper = cat.upper()
        mat_rate  = float(maternity_rates.get(cat, maternity_rates.get('A', 0)) or 0)
        categories_data[cat_upper] = {
            'brackets':      brackets,
            'maternity_rate': mat_rate,
        }

    if not categories_data:
        return jsonify({'error': 'No rate data provided'}), 400

    # ── Maternity validation for Healthx / Healthxclusive ─────────────────────
    plan_lower = form_data.get('plan', '').lower()
    if plan_lower in ('healthx', 'healthxclusive'):
        missing_mat = [cat for cat, data in categories_data.items()
                       if not float(data.get('maternity_rate') or 0)]
        if missing_mat:
            cats_str = ', '.join(sorted(missing_mat))
            return jsonify({
                'error': (
                    f'Maternity premium is required for {plan_lower.capitalize()} but is missing '
                    f'for category/categories: {cats_str}. '
                    f'Please enter the maternity surcharge in the rate panel before calculating.'
                )
            }), 400

    # Ensure all census categories have a rate (fallback)
    first_cat = sorted(categories_data.keys())[0]
    for m in members:
        c = m['category'].upper()
        if c not in categories_data:
            categories_data[c] = categories_data[first_cat]

    members_data, totals = calculate_premiums(members, categories_data)
    company_name         = form_data.get('company_name', 'Company')

    out_token = str(uuid.uuid4())
    _session_set(out_token, {
        'company_name':   company_name,
        'members_data':   members_data,
        'verified_rates': verified_rates,
        'maternity_rates': maternity_rates,
        'form_data':      form_data,
        'quote_totals':   quote_totals,
        'totals':         totals,
        'plan':           stored.get('plan', ''),
        'start_date':     start_date,
        'age_method':     age_method,
    })

    q_premium = float(quote_totals.get('total_premium', 0) or 0)
    c_premium = totals['total_net'] + totals['total_maternity']
    diff      = c_premium - q_premium
    is_match  = abs(diff) < 20.0
    q_members = int(quote_totals.get('members', 0) or 0)

    diff_items = []
    if not is_match and q_premium > 0:
        diff_items.append({'label': 'Net Premium', 'quote': q_premium, 'calc': c_premium, 'diff': diff})
    if q_members and q_members != len(members_data):
        diff_items.append({'label': 'Member Count', 'quote': q_members, 'calc': len(members_data),
                           'diff': len(members_data) - q_members, 'is_count': True})

    return jsonify({
        'success':      True,
        'out_token':    out_token,
        'totals':       totals,
        'member_count': len(members_data),
        'members': [
            {'name': m['name'], 'dob': str(m['dob']),
             'gender': m['gender'], 'category': m['category'],
             'relation': m['relation']}
            for m in members_data
        ],
        'reconciliation': {
            'quote_premium': q_premium,
            'calc_premium':  c_premium,
            'difference':    diff,
            'match':         is_match,
            'items':         diff_items,
        }
    })

@app.route('/api/finalize_summary', methods=['POST'])
def api_finalize_summary():
    body            = request.json or {}
    out_token       = body.get('out_token', '')
    loading_members = body.get('loading_members', [])
    has_lsb         = bool(body.get('has_lsb', False))

    stored = _session_get(out_token)
    if not stored or 'members_data' not in stored:
        return jsonify({'error': 'Session expired. Please recalculate.'}), 400

    # Merge any overrides from the finalize call into form_data
    fd = dict(stored['form_data'])
    if body.get('inception_payment'):
        fd['inception_payment'] = body['inception_payment']
    if body.get('endorsement_freq'):
        fd['endorsement_freq'] = body['endorsement_freq']
    fd['has_lsb'] = has_lsb

    quote_totals = stored.get('quote_totals', {})
    totals       = stored.get('totals', {})

    # Pass recon_note from finalize request into form_data for save_policy
    if body.get('recon_note') is not None:
        fd['recon_note'] = body['recon_note']

    census_diff_summary = stored.get('census_diff_summary', '')

    try:
        wb = make_combined_excel(
            fd,
            stored['members_data'],
            stored['verified_rates'],
            stored['maternity_rates'],
            loading_members,
            has_lsb,
            totals,
            quote_totals,
            quoted_totals_calc=None,
            quoted_member_count=None,
            census_diff_summary=census_diff_summary,
        )
    except Exception as e:
        return jsonify({'error': f'Excel generation error: {str(e)}'}), 500

    buf = io.BytesIO()
    wb.save(buf)

    prd_token = str(uuid.uuid4())
    _session_set(prd_token, {
        'prd_bytes':    buf.getvalue(),
        'company_name': stored.get('company_name', 'Company'),
    })

    # ── Persist to Supabase ───────────────────────────────────────────────────
    policy_id = save_policy(
        fd,
        stored['members_data'],
        stored['verified_rates'],
        stored['maternity_rates'],
        loading_members,
        totals,
        quote_totals,
    )

    return jsonify({'prd_token': prd_token, 'policy_id': policy_id})


@app.route('/download/<token>/<file_type>')
def download(token, file_type):
    stored = _session_get(token)
    if not stored:
        return "File not found or expired", 404
    company = re.sub(r'[^\w\s-]', '', stored.get('company_name', 'Company')).strip()
    if file_type == 'prd':
        data, filename = stored['prd_bytes'], f"Premium Summary - {company}.xlsx"
    else:
        return "Unknown file type", 400
    return send_file(io.BytesIO(data), as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Policy Database Pages ─────────────────────────────────────────────────────

@app.route('/policies')
def policies_page():
    return render_template('policies.html')


@app.route('/policies/<int:pid>')
def policy_detail_page(pid):
    return render_template('policy_detail.html', policy_id=pid)


@app.route('/dashboard')
def dashboard_page():
    return render_template('dashboard.html')


# ── Policy Database API ───────────────────────────────────────────────────────

@app.route('/api/policies')
def api_policies():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    search    = request.args.get('search', '').strip()
    broker    = request.args.get('broker', '').strip()
    rm        = request.args.get('rm', '').strip()
    plan      = request.args.get('plan', '').strip()
    ptype     = request.args.get('plan_type', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    page      = max(1, int(request.args.get('page', 1)))
    per_page  = 20
    offset    = (page - 1) * per_page

    try:
        q = supa.table('policies').select(
            'id,created_at,company_name,broker,rm_person,plan,plan_type,start_date,member_count,grand_total,recon_match',
            count='exact'
        )
        if search:
            q = q.ilike('company_name', f'%{search}%')
        if broker:
            q = q.eq('broker', broker)
        if rm:
            q = q.eq('rm_person', rm)
        if plan:
            q = q.eq('plan', plan)
        if ptype:
            q = q.eq('plan_type', ptype)
        if date_from:
            q = q.gte('start_date', date_from)
        if date_to:
            q = q.lte('start_date', date_to)
        q = q.order('created_at', desc=True).range(offset, offset + per_page - 1)
        res   = q.execute()
        rows  = res.data or []
        total = res.count or 0
        return jsonify({'rows': rows, 'total': total, 'page': page, 'per_page': per_page})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Shared helper: recalculate members + totals from stored rates ─────────────

def _recalculate_policy(supa, policy_id):
    """Re-apply stored bracket rates to all members, then update policy totals."""
    # Fetch categories + brackets
    cat_res = supa.table('policy_categories').select('*').eq('policy_id', policy_id).execute()
    categories_data = {}
    for cat in (cat_res.data or []):
        br_res = supa.table('policy_brackets').select('*') \
                     .eq('category_id', cat['id']).order('age_lo').execute()
        categories_data[cat['category'].upper()] = {
            'brackets': [
                {'age_lo': b['age_lo'], 'age_hi': b['age_hi'],
                 'male': b['male_rate'], 'female': b['female_rate'], 'label': b['label']}
                for b in (br_res.data or [])
            ],
            'maternity_rate': float(cat.get('maternity_rate') or 0),
        }

    # Fetch members
    mem_res = supa.table('policy_members').select('*').eq('policy_id', policy_id).execute()
    members = mem_res.data or []

    total_net = 0.0
    total_mat = 0.0
    for m in members:
        rate, bracket_label, _ = get_member_rate(m, categories_data)
        cat      = (m.get('category') or 'A').upper()
        mat_rate = float(categories_data.get(cat, {}).get('maternity_rate') or 0)
        emirate  = m.get('emirate') or 'Dubai'
        mat_age_max = MAT_AGE_MAX_AUH if 'abu dhabi' in emirate.lower() else MAT_AGE_MAX_DXB
        gender   = (m.get('gender') or '').lower()
        marital  = (m.get('marital_status') or '').lower()
        age      = m.get('age_alb') or 0
        maternity = mat_rate if (
            gender.startswith('f') and marital.startswith('m')
            and MAT_AGE_MIN <= age <= mat_age_max
        ) else 0.0
        gross_load = float(m.get('gross_loading') or 0)
        final_prem = float(rate) + maternity + gross_load

        supa.table('policy_members').update({
            'base_premium':      float(rate),
            'maternity_premium': maternity,
            'final_premium':     final_prem,
            'age_bracket':       bracket_label or m.get('age_bracket', ''),
        }).eq('id', m['id']).execute()

        total_net += float(rate)
        total_mat += maternity

    bas_total  = BASMAH_FEE * len(members)
    subtotal   = total_net + total_mat + bas_total
    vat        = subtotal * VAT_RATE
    grand_total = subtotal + vat

    supa.table('policies').update({
        'total_net':       total_net,
        'total_maternity': total_mat,
        'total_basmah':    bas_total,
        'subtotal':        subtotal,
        'vat':             vat,
        'grand_total':     grand_total,
    }).eq('id', policy_id).execute()


# ── Edit rates ────────────────────────────────────────────────────────────────

@app.route('/api/policies/<int:pid>/rates', methods=['PUT'])
def api_policy_rates_edit(pid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    data       = request.get_json(force=True) or {}
    categories = data.get('categories', [])
    try:
        for cat in categories:
            cat_id = cat.get('id')
            if not cat_id:
                continue
            supa.table('policy_categories').update({
                'maternity_rate': float(cat.get('maternity_rate') or 0),
            }).eq('id', cat_id).execute()

            for br in cat.get('brackets', []):
                br_id = br.get('id')
                if not br_id:
                    continue
                supa.table('policy_brackets').update({
                    'male_rate':   float(br.get('male_rate')   or 0),
                    'female_rate': float(br.get('female_rate') or 0),
                }).eq('id', br_id).execute()

        _recalculate_policy(supa, pid)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Replace census ────────────────────────────────────────────────────────────

@app.route('/api/policies/<int:pid>/census', methods=['POST'])
def api_policy_census_replace(pid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503

    census_file = request.files.get('census')
    if not census_file:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        pol_res = supa.table('policies').select('start_date,plan').eq('id', pid).execute()
        if not pol_res.data:
            return jsonify({'error': 'Policy not found'}), 404
        pol        = pol_res.data[0]
        start_date = str(pol.get('start_date') or '')
        plan_lower = (pol.get('plan') or '').lower()
        age_method = 'anb' if plan_lower == 'healthx' else 'alb'

        file_bytes = census_file.read()
        filename   = census_file.filename or 'census.xlsx'

        # Parse new census
        members_raw, _ = parse_census(file_bytes, filename, start_date, age_method)

        # Fetch stored rates
        cat_res = supa.table('policy_categories').select('*').eq('policy_id', pid).execute()
        categories_data = {}
        for cat in (cat_res.data or []):
            br_res = supa.table('policy_brackets').select('*') \
                         .eq('category_id', cat['id']).order('age_lo').execute()
            categories_data[cat['category'].upper()] = {
                'brackets': [
                    {'age_lo': b['age_lo'], 'age_hi': b['age_hi'],
                     'male': b['male_rate'], 'female': b['female_rate'], 'label': b['label']}
                    for b in (br_res.data or [])
                ],
                'maternity_rate': float(cat.get('maternity_rate') or 0),
            }

        # Calculate premiums using stored rates — returns (members_list, totals_dict)
        members_calc, _ = calculate_premiums(members_raw, categories_data)

        if not members_calc:
            return jsonify({'error': 'No valid members found in census file'}), 400

        # Build ALL new rows FIRST — no DB writes yet (atomic swap safety)
        mem_rows = []
        for i, m in enumerate(members_calc):
            gross_load = float(m.get('gross_loading') or 0)
            mem_rows.append({
                'policy_id':        pid,
                'member_no':        i + 1,
                'name':             m.get('name', ''),
                'dob':              _parse_dob(str(m.get('dob', ''))),
                'gender':           m.get('gender', ''),
                'marital_status':   m.get('marital_status', ''),
                'relation':         m.get('relation', ''),
                'category':         m.get('category', ''),
                'age_alb':          m.get('age_alb'),
                'emirate':          m.get('emirate', ''),
                'age_bracket':      m.get('age_bracket', ''),
                'base_premium':     m['base_premium'],
                'maternity_premium': m['maternity_premium'],
                'gross_loading':    gross_load,
                'final_premium':    m['base_premium'] + m['maternity_premium'] + gross_load,
            })

        # Only NOW do the destructive swap — existing members are safe until this point
        supa.table('policy_members').delete().eq('policy_id', pid).execute()
        for i in range(0, len(mem_rows), 500):
            supa.table('policy_members').insert(mem_rows[i:i + 500]).execute()

        # Recalculate and update policy totals
        total_net  = sum(m['base_premium']      for m in members_calc)
        total_mat  = sum(m['maternity_premium'] for m in members_calc)
        bas_total  = BASMAH_FEE * len(members_calc)
        subtotal   = total_net + total_mat + bas_total
        vat        = subtotal * VAT_RATE
        grand_total = subtotal + vat

        supa.table('policies').update({
            'member_count':    len(members_calc),
            'total_net':       total_net,
            'total_maternity': total_mat,
            'total_basmah':    bas_total,
            'subtotal':        subtotal,
            'vat':             vat,
            'grand_total':     grand_total,
        }).eq('id', pid).execute()

        return jsonify({'ok': True, 'member_count': len(members_calc)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Regenerate / Export premium summary Excel ────────────────────────────────

@app.route('/api/policies/<int:pid>/export')
def api_policy_export(pid):
    """Rebuild and stream the Premium Summary Excel from stored DB data."""
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        # ── 1. Fetch policy row ──────────────────────────────────────────────
        pol_res = supa.table('policies').select('*').eq('id', pid).execute()
        if not pol_res.data:
            return jsonify({'error': 'Policy not found'}), 404
        pol = pol_res.data[0]

        # ── 2. Fetch members ─────────────────────────────────────────────────
        mem_res = supa.table('policy_members').select('*') \
                      .eq('policy_id', pid).order('member_no').execute()
        members_db = mem_res.data or []

        # ── 3. Fetch categories + brackets ──────────────────────────────────
        cat_res = supa.table('policy_categories').select('*') \
                      .eq('policy_id', pid).order('category').execute()
        verified_rates  = {}
        maternity_rates = {}
        for cat in (cat_res.data or []):
            br_res = supa.table('policy_brackets').select('*') \
                         .eq('category_id', cat['id']).order('age_lo').execute()
            cat_letter = cat['category'].upper()
            verified_rates[cat_letter] = [
                {
                    'age_lo': b['age_lo'],
                    'age_hi': b['age_hi'],
                    'male':   float(b['male_rate']   or 0),
                    'female': float(b['female_rate'] or 0),
                    'label':  b['label'],
                }
                for b in (br_res.data or [])
            ]
            maternity_rates[cat_letter] = float(cat.get('maternity_rate') or 0)

        # ── 4. Reconstruct form_data ─────────────────────────────────────────
        form_data = {
            'company_name':     pol.get('company_name', ''),
            'broker':           pol.get('broker', ''),
            'underwriter':      pol.get('underwriter', ''),
            'rm_person':        pol.get('rm_person', ''),
            'plan':             pol.get('plan', ''),
            'plan_type':        pol.get('plan_type', ''),
            'start_date':       str(pol.get('start_date') or ''),
            'inception_payment': pol.get('inception_payment') or 'Annual',
            'endorsement_freq': pol.get('endorsement_freq') or 'Monthly',
            'has_lsb':          bool(pol.get('has_lsb', False)),
            'rm_broker':        _f(pol.get('rm_broker')),
            'rm_insurer':       _f(pol.get('rm_insurer')),
            'rm_wellx':         _f(pol.get('rm_wellx')),
            'rm_tpa':           _f(pol.get('rm_tpa')),
            'rm_insurance_tax': _f(pol.get('rm_insurance_tax')),
        }
        has_lsb = form_data['has_lsb']

        # ── 5. Reconstruct members_data + loading_members ────────────────────
        members_data    = []
        loading_members = []
        for m in members_db:
            members_data.append({
                'no':               m.get('member_no') or 0,
                'name':             m.get('name', ''),
                'dob':              m.get('dob', '') or '',
                'gender':           m.get('gender', ''),
                'category':         m.get('category', ''),
                'relation':         m.get('relation', ''),
                'marital_status':   m.get('marital_status', ''),
                'emirate':          m.get('emirate', ''),
                'age_alb':          m.get('age_alb') or 0,
                'age_bracket':      m.get('age_bracket', ''),
                'base_premium':     float(m.get('base_premium')     or 0),
                'maternity_premium': float(m.get('maternity_premium') or 0),
                'error':            '',
            })
            gross_load = float(m.get('gross_loading') or 0)
            if gross_load > 0:
                loading_members.append({
                    'name':         m.get('name', ''),
                    'dob':          m.get('dob', '') or '',
                    'gender':       m.get('gender', ''),
                    'category':     m.get('category', ''),
                    'relation':     m.get('relation', ''),
                    'gross_loading': gross_load,
                })

        # ── 6. Reconstruct totals + quote_totals ─────────────────────────────
        totals = {
            'total_net':       _f(pol.get('total_net')),
            'total_maternity': _f(pol.get('total_maternity')),
            'total_basmah':    _f(pol.get('total_basmah')),
            'subtotal':        _f(pol.get('subtotal')),
            'vat':             _f(pol.get('vat')),
            'grand_total':     _f(pol.get('grand_total')),
            'member_count':    int(pol.get('member_count') or 0),
        }
        quote_totals = {
            'total_premium': _f(pol.get('confirmed_quote')),
            'members':       int(pol.get('quoted_members') or 0),
            'grand_total':   _f(pol.get('quote_grand_total')),
        }

        # ── 7. Generate Excel ────────────────────────────────────────────────
        wb = make_combined_excel(
            form_data, members_data, verified_rates, maternity_rates,
            loading_members, has_lsb, totals, quote_totals,
            hide_commissions=True,
        )
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        company  = re.sub(r'[^\w\s-]', '', pol.get('company_name', 'Company')).strip()
        filename = f"Premium Summary - {company}.xlsx"
        return send_file(
            buf, as_attachment=True, download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Shared: resync policy totals from current member rows ─────────────────────

def _resync_policy_totals(supa, policy_id):
    """Sum existing member premiums and update the policy record totals."""
    mem_res    = supa.table('policy_members').select(
        'base_premium,maternity_premium,gross_loading'
    ).eq('policy_id', policy_id).execute()
    members    = mem_res.data or []
    total_net  = sum(float(m.get('base_premium')      or 0) for m in members)
    total_mat  = sum(float(m.get('maternity_premium') or 0) for m in members)
    bas_total  = BASMAH_FEE * len(members)
    subtotal   = total_net + total_mat + bas_total
    vat        = subtotal * VAT_RATE
    supa.table('policies').update({
        'member_count':    len(members),
        'total_net':       total_net,
        'total_maternity': total_mat,
        'total_basmah':    bas_total,
        'subtotal':        subtotal,
        'vat':             vat,
        'grand_total':     subtotal + vat,
    }).eq('id', policy_id).execute()


# ── Shared: calculate one member's premium from stored rates ──────────────────

def _calc_member_prem(member_data, supa, policy_id):
    """
    Given a member dict (name, dob, gender, marital_status, relation,
    category, emirate, age_alb), return (base_premium, maternity_premium,
    age_bracket) using this policy's stored bracket rates.
    """
    cat_res = supa.table('policy_categories').select('*').eq('policy_id', policy_id).execute()
    categories_data = {}
    for cat in (cat_res.data or []):
        br_res = supa.table('policy_brackets').select('*') \
                     .eq('category_id', cat['id']).order('age_lo').execute()
        categories_data[cat['category'].upper()] = {
            'brackets': [
                {'age_lo': b['age_lo'], 'age_hi': b['age_hi'],
                 'male': b['male_rate'], 'female': b['female_rate'], 'label': b['label']}
                for b in (br_res.data or [])
            ],
            'maternity_rate': float(cat.get('maternity_rate') or 0),
        }

    rate, bracket_label, _ = get_member_rate(member_data, categories_data)
    cat_key     = (member_data.get('category') or '').upper()
    mat_rate    = categories_data.get(cat_key, {}).get('maternity_rate', 0)
    emirate     = member_data.get('emirate') or 'Dubai'
    mat_age_max = MAT_AGE_MAX_AUH if 'abu dhabi' in emirate.lower() else MAT_AGE_MAX_DXB
    gender      = (member_data.get('gender') or '').lower()
    marital     = (member_data.get('marital_status') or '').lower()
    age         = member_data.get('age_alb') or 0
    maternity   = float(mat_rate or 0) if (
        gender.startswith('f') and marital.startswith('m')
        and MAT_AGE_MIN <= age <= mat_age_max
    ) else 0.0
    return float(rate), maternity, bracket_label or ''


# ── Member CRUD ───────────────────────────────────────────────────────────────

def _member_from_body(body, supa, policy_id, start_date, plan_lower):
    """Parse request body into a member dict with calculated age & premiums."""
    dob_str = (body.get('dob') or '').strip()
    if not dob_str:
        raise ValueError('Date of birth is required')

    # Parse DOB → date object
    dob_date = None
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%d-%b-%Y'):
        try:
            dob_date = datetime.strptime(dob_str, fmt).date()
            break
        except ValueError:
            continue
    if not dob_date:
        raise ValueError(f'Cannot parse date of birth: {dob_str}')

    # Calculate age
    age_method = 'anb' if plan_lower == 'healthx' else 'alb'
    age_fn     = calculate_anb if age_method == 'anb' else calculate_alb
    try:
        sd = datetime.strptime(str(start_date), '%Y-%m-%d').date()
    except Exception:
        sd = date.today()
    age_alb = age_fn(dob_date, sd)

    member = {
        'name':           (body.get('name') or '').strip(),
        'dob':            dob_date,
        'gender':         body.get('gender', 'Male'),
        'marital_status': body.get('marital_status', 'Single'),
        'relation':       body.get('relation', 'Employee'),
        'category':       (body.get('category') or 'A').upper(),
        'emirate':        body.get('emirate', 'Dubai'),
        'age_alb':        age_alb,
    }

    base_prem, mat_prem, bracket_label = _calc_member_prem(member, supa, policy_id)
    gross_load  = float(body.get('gross_loading') or 0)
    final_prem  = base_prem + mat_prem + gross_load

    return {
        **member,
        'dob':              dob_date.strftime('%Y-%m-%d'),
        'age_bracket':      bracket_label,
        'base_premium':     base_prem,
        'maternity_premium': mat_prem,
        'gross_loading':    gross_load,
        'final_premium':    final_prem,
    }


@app.route('/api/policies/<int:pid>/members', methods=['POST'])
def api_member_add(pid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        pol_res = supa.table('policies').select('start_date,plan').eq('id', pid).execute()
        if not pol_res.data:
            return jsonify({'error': 'Policy not found'}), 404
        pol        = pol_res.data[0]
        start_date = str(pol.get('start_date') or '')
        plan_lower = (pol.get('plan') or '').lower()

        body   = request.get_json(force=True) or {}
        member = _member_from_body(body, supa, pid, start_date, plan_lower)

        # Get next member_no
        cnt_res  = supa.table('policy_members').select('member_no') \
                       .eq('policy_id', pid).order('member_no', desc=True).limit(1).execute()
        next_no  = (cnt_res.data[0]['member_no'] + 1) if cnt_res.data else 1
        member['policy_id']  = pid
        member['member_no']  = next_no

        res = supa.table('policy_members').insert(member).execute()
        _resync_policy_totals(supa, pid)
        return jsonify({'ok': True, 'member': res.data[0] if res.data else {}}), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/policies/<int:pid>/members/<int:mid>', methods=['PUT'])
def api_member_edit(pid, mid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        pol_res = supa.table('policies').select('start_date,plan').eq('id', pid).execute()
        if not pol_res.data:
            return jsonify({'error': 'Policy not found'}), 404
        pol        = pol_res.data[0]
        start_date = str(pol.get('start_date') or '')
        plan_lower = (pol.get('plan') or '').lower()

        body   = request.get_json(force=True) or {}
        member = _member_from_body(body, supa, pid, start_date, plan_lower)

        res = supa.table('policy_members').update(member).eq('id', mid).eq('policy_id', pid).execute()
        if not res.data:
            return jsonify({'error': 'Member not found'}), 404
        _resync_policy_totals(supa, pid)
        return jsonify({'ok': True, 'member': res.data[0]})
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/policies/<int:pid>/members/<int:mid>', methods=['DELETE'])
def api_member_delete(pid, mid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        res = supa.table('policy_members').delete().eq('id', mid).eq('policy_id', pid).execute()
        if not res.data:
            return jsonify({'error': 'Member not found'}), 404
        _resync_policy_totals(supa, pid)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/policies/<int:pid>', methods=['PUT'])
def api_policy_edit(pid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    data = request.get_json(force=True) or {}
    # Whitelist of editable fields
    allowed = {
        'company_name', 'broker', 'rm_person', 'underwriter',
        'plan', 'plan_type', 'start_date', 'confirmation_date', 'recon_note',
        'inception_payment', 'endorsement_freq', 'has_lsb',
        'rm_broker', 'rm_insurer', 'rm_wellx', 'rm_tpa', 'rm_insurance_tax',
    }
    update = {k: v for k, v in data.items() if k in allowed}
    if not update:
        return jsonify({'error': 'No valid fields to update'}), 400
    # Coerce types
    for f in ('rm_broker', 'rm_insurer', 'rm_wellx', 'rm_tpa', 'rm_insurance_tax'):
        if f in update:
            try:
                update[f] = float(update[f])
            except (TypeError, ValueError):
                update[f] = None
    if 'has_lsb' in update:
        update['has_lsb'] = bool(update['has_lsb'])
    try:
        res = supa.table('policies').update(update).eq('id', pid).execute()
        if not res.data:
            return jsonify({'error': 'Not found'}), 404
        return jsonify({'ok': True, 'policy': res.data[0]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/policies/<int:pid>', methods=['DELETE'])
def api_policy_delete(pid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        res = supa.table('policies').delete().eq('id', pid).execute()
        if not res.data:
            return jsonify({'error': 'Not found'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/policies/bulk_delete', methods=['POST'])
def api_policies_bulk_delete():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs provided'}), 400
    try:
        supa.table('policies').delete().in_('id', ids).execute()
        return jsonify({'ok': True, 'deleted': len(ids)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/policies/<int:pid>')
def api_policy_detail(pid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        pol_res = supa.table('policies').select('*').eq('id', pid).execute()
        if not pol_res.data:
            return jsonify({'error': 'Not found'}), 404
        pol = pol_res.data[0]

        mem_res = supa.table('policy_members').select('*').eq('policy_id', pid).order('member_no').execute()
        members = mem_res.data or []

        cat_res = supa.table('policy_categories').select('*').eq('policy_id', pid).order('category').execute()
        categories = []
        for cat in (cat_res.data or []):
            br_res = supa.table('policy_brackets').select('*').eq('category_id', cat['id']).order('age_lo').execute()
            cat['brackets'] = br_res.data or []
            categories.append(cat)

        return jsonify({'policy': pol, 'members': members, 'categories': categories})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _dash_filter(q, args):
    """Apply shared dashboard filters to a supabase query."""
    from collections import defaultdict as _dd
    plan      = args.get('plan', '').strip()
    plan_type = args.get('plan_type', '').strip()
    rm        = args.get('rm', '').strip()
    broker    = args.get('broker', '').strip()
    if plan:
        q = q.eq('plan', plan)
    if plan_type:
        q = q.eq('plan_type', plan_type)
    if rm:
        q = q.eq('rm_person', rm)
    if broker:
        q = q.eq('broker', broker)
    return q


@app.route('/api/dashboard/monthly')
def api_dashboard_monthly():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        from collections import defaultdict
        q = supa.table('policies').select(
            'created_at,plan,plan_type,total_net,total_maternity,grand_total,member_count'
        )
        q = _dash_filter(q, request.args)
        rows = q.execute().data or []

        # Per-month summary (for the table)
        monthly = defaultdict(lambda: {'policy_count': 0, 'gross_premium': 0.0, 'grand_total': 0.0, 'member_count': 0})
        # Per-month + plan + plan_type breakdown (for the chart)
        breakdown = defaultdict(lambda: {'policy_count': 0, 'gross_premium': 0.0})

        for r in rows:
            m     = (r.get('created_at') or '')[:7]   # YYYY-MM
            plan  = r.get('plan')      or 'Unknown'
            ptype = r.get('plan_type') or 'Unknown'
            gross = float(r.get('total_net') or 0) + float(r.get('total_maternity') or 0)
            grand = float(r.get('grand_total') or 0)
            mems  = int(r.get('member_count') or 0)

            monthly[m]['policy_count']  += 1
            monthly[m]['gross_premium'] += gross
            monthly[m]['grand_total']   += grand
            monthly[m]['member_count']  += mems

            key = f'{m}|{plan}|{ptype}'
            breakdown[key]['policy_count']  += 1
            breakdown[key]['gross_premium'] += gross

        return jsonify({
            'monthly': [{'month': k, **v} for k, v in sorted(monthly.items())],
            'breakdown': [
                {'month': k.split('|')[0], 'plan': k.split('|')[1], 'plan_type': k.split('|')[2], **v}
                for k, v in sorted(breakdown.items())
            ],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/brokers')
def api_dashboard_brokers():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        from collections import defaultdict
        q = supa.table('policies').select('broker,member_count,total_net,total_maternity,grand_total')
        q = _dash_filter(q, request.args)
        rows = q.execute().data or []
        buckets = defaultdict(lambda: {'policy_count': 0, 'total_members': 0, 'gross_premium': 0.0, 'total_grand': 0.0})
        for r in rows:
            b = r.get('broker') or 'Unknown'
            buckets[b]['policy_count']  += 1
            buckets[b]['total_members'] += int(r.get('member_count') or 0)
            buckets[b]['gross_premium'] += float(r.get('total_net') or 0) + float(r.get('total_maternity') or 0)
            buckets[b]['total_grand']   += float(r.get('grand_total') or 0)
        result = []
        for k, v in buckets.items():
            v['broker'] = k
            v['avg_policy_size'] = round(v['total_grand'] / v['policy_count'], 2) if v['policy_count'] else 0
            result.append(v)
        result.sort(key=lambda x: x['total_grand'], reverse=True)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/rm')
def api_dashboard_rm():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        from collections import defaultdict
        q = supa.table('policies').select('rm_person,created_at,grand_total')
        q = _dash_filter(q, request.args)
        rows = q.execute().data or []
        buckets = defaultdict(lambda: {'policy_count': 0, 'actual_amount': 0.0})
        for r in rows:
            rm_name = r.get('rm_person') or 'Unassigned'
            dt = r.get('created_at', '')[:10]
            try:
                d = datetime.strptime(dt, '%Y-%m-%d')
                key = (rm_name, d.year, d.month)
            except Exception:
                continue
            buckets[key]['policy_count']  += 1
            buckets[key]['actual_amount'] += float(r.get('grand_total') or 0)
        actuals = [{'rm_name': k[0], 'year': k[1], 'month': k[2], **v} for k, v in buckets.items()]
        actuals.sort(key=lambda x: (x['rm_name'], x['year'], x['month']))
        tgt_res = supa.table('rm_targets').select('*').order('rm_name').execute()
        return jsonify({'actuals': actuals, 'targets': tgt_res.data or []})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/rm_targets', methods=['POST'])
def api_upsert_rm_target():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    body          = request.json or {}
    rm_name       = body.get('rm_name', '').strip()
    year          = int(body.get('year', 0))
    month         = int(body.get('month', 0))
    target_amount = float(body.get('target_amount', 0) or 0)
    if not rm_name or not year or not month:
        return jsonify({'error': 'rm_name, year, month are required'}), 400
    try:
        res = supa.table('rm_targets').upsert({
            'rm_name': rm_name, 'year': year, 'month': month, 'target_amount': target_amount
        }, on_conflict='rm_name,year,month').execute()
        return jsonify({'ok': True, 'id': res.data[0]['id'] if res.data else None})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Logo helper ───────────────────────────────────────────────────────────────
_WELLX_LOGO_PATH   = os.path.join(_ROOT, 'Style', 'wellx-logo-light-bg_de94776e.png')
_POWERED_LOGO_PATH = BRAND_LOGO_PATH   # "Powered by Wellx Labs no background.png"

def _xl_add_logo(ws, cell='A1', path=None, w=130, h=38):
    """Insert Wellx logo image into ws at cell. Silently skips if Pillow unavailable."""
    try:
        from openpyxl.drawing.image import Image as _XLImg
        p = path or _WELLX_LOGO_PATH
        if os.path.isfile(p):
            img = _XLImg(p)
            img.width  = w
            img.height = h
            ws.add_image(img, cell)
    except Exception:
        pass   # Pillow not installed or file missing — degrade gracefully


def _xl_powered_row(ws, row, col=1, last_col=8):
    """Write 'Powered by Wellx Labs' text in the footer row."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=col,
                value=f'Powered by Wellx Labs  ·  Generated {datetime.now().strftime("%d %b %Y %H:%M")}')
    c.font      = Font(name='Inter', size=8, italic=True, color='9aa5b4')
    c.alignment = Alignment(horizontal='center', vertical='center')


def _make_report_header(ws, title, subtitle='', logo=True, col_count=8):
    """Write a branded header (logo + title) into rows 1-2 of ws. Returns next free row."""
    from openpyxl.styles import PatternFill as _PF
    PRI = '003780'
    ws.row_dimensions[1].height = 48
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(name='Raleway', bold=True, size=16, color='FFFFFF')
    c.fill      = _PF('solid', fgColor=PRI)
    c.alignment = Alignment(horizontal='center', vertical='center')
    if logo:
        _xl_add_logo(ws, 'A1', _WELLX_LOGO_PATH, w=110, h=36)

    if subtitle:
        ws.row_dimensions[2].height = 20
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font      = Font(name='Inter', size=9.5, italic=True, color='6b7280')
        s.fill      = _PF('solid', fgColor='f8fafc')
        s.alignment = Alignment(horizontal='center', vertical='center')
        return 3
    return 2


# ── Take Rate API ─────────────────────────────────────────────────────────────
@app.route('/api/dashboard/take_rate')
def api_dashboard_take_rate():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        from collections import defaultdict
        q = supa.table('policies').select(
            'plan,total_net,total_maternity,'
            'rm_broker,rm_insurer,rm_wellx,rm_tpa,rm_insurance_tax'
        )
        q = _dash_filter(q, request.args)
        rows = q.execute().data or []

        buckets = defaultdict(lambda: {'premium': 0.0, 'commission': 0.0, 'count': 0})
        for r in rows:
            plan     = r.get('plan') or 'Unknown'
            gross    = float(r.get('total_net') or 0) + float(r.get('total_maternity') or 0)
            comm_pct = (
                float(r.get('rm_broker', 0) or 0) +
                float(r.get('rm_insurer', 0) or 0) +
                float(r.get('rm_wellx', 0) or 0) +
                float(r.get('rm_tpa', 0) or 0) +
                float(r.get('rm_insurance_tax', 0) or 0)
            )
            comm_amt = gross * comm_pct / 100.0
            buckets[plan]['premium']    += gross
            buckets[plan]['commission'] += comm_amt
            buckets[plan]['count']      += 1

        rows_out = []
        total_prem = total_comm = 0.0
        for plan, v in sorted(buckets.items()):
            prem  = round(v['premium'],    2)
            comm  = round(v['commission'], 2)
            rate  = round(comm / prem * 100, 2) if prem else 0.0
            rows_out.append({'plan': plan, 'policy_count': v['count'],
                             'total_premium': prem, 'commission_amount': comm,
                             'take_rate_pct': rate})
            total_prem += prem
            total_comm += comm

        total_rate = round(total_comm / total_prem * 100, 2) if total_prem else 0.0
        return jsonify({
            'plans': rows_out,
            'total': {
                'total_premium':     round(total_prem, 2),
                'commission_amount': round(total_comm, 2),
                'take_rate_pct':     total_rate,
            },
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Dashboard Excel Export ────────────────────────────────────────────────────
@app.route('/api/dashboard/export')
def api_dashboard_export():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        from collections import defaultdict
        today_str = datetime.now().strftime('%d %b %Y')
        args = request.args

        # ── Fetch data ──────────────────────────────────────────────────────
        q_monthly = supa.table('policies').select(
            'created_at,plan,plan_type,total_net,total_maternity,grand_total,member_count'
        )
        q_monthly = _dash_filter(q_monthly, args)
        all_rows  = q_monthly.execute().data or []

        q_tr = supa.table('policies').select(
            'plan,total_net,total_maternity,'
            'rm_broker,rm_insurer,rm_wellx,rm_tpa,rm_insurance_tax'
        )
        q_tr = _dash_filter(q_tr, args)
        tr_rows = q_tr.execute().data or []

        q_br = supa.table('policies').select(
            'broker,member_count,total_net,total_maternity,grand_total'
        )
        q_br = _dash_filter(q_br, args)
        br_rows = q_br.execute().data or []

        # ── Build workbook ──────────────────────────────────────────────────
        from openpyxl import Workbook as _WB
        from openpyxl.styles import PatternFill as _PF
        wb  = _WB()
        ws  = wb.active
        ws.title = 'Wellx Policies Report'
        PRI = '003780'; ORG = 'E86F2C'; LGT = 'EFF4FB'; HDR = 'DBEAFE'

        col_widths = [18, 12, 14, 18, 16, 14, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = _make_report_header(ws,
            f'Wellx Policies Report — as of {today_str}',
            subtitle='Auto-generated · Wellx Labs',
            col_count=7) + 1

        def hdr_cell(r, c, val, bg=PRI, fg='FFFFFF', bold=True, size=9.5):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = Font(name='Inter', bold=bold, size=size, color=fg)
            cell.fill      = _PF('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            return cell

        def val_cell(r, c, val, bg='FFFFFF', bold=False, num_fmt=None, halign='right'):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = Font(name='Inter', size=9.5, bold=bold, color='1e293b')
            cell.fill      = _PF('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal=halign, vertical='center')
            if num_fmt:
                cell.number_format = num_fmt
            return cell

        # ── Monthly summary table ───────────────────────────────────────────
        monthly = defaultdict(lambda: {'policy_count': 0, 'gross_premium': 0.0, 'grand_total': 0.0, 'member_count': 0})
        for r in all_rows:
            m     = (r.get('created_at') or '')[:7]
            gross = float(r.get('total_net') or 0) + float(r.get('total_maternity') or 0)
            monthly[m]['policy_count']  += 1
            monthly[m]['gross_premium'] += gross
            monthly[m]['grand_total']   += float(r.get('grand_total') or 0)
            monthly[m]['member_count']  += int(r.get('member_count') or 0)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value='Monthly Summary').font = Font(name='Raleway', bold=True, size=11, color=PRI)
        row += 1
        for c, h in enumerate(['Month', 'Policies', 'Members', 'Gross Premium', 'Grand Total'], 1):
            hdr_cell(row, c, h)
        row += 1
        sg = st = sm = 0.0; sc = 0
        for month_key in sorted(monthly):
            v   = monthly[month_key]
            alt = _PF('solid', fgColor=LGT if row % 2 == 0 else 'FFFFFF')
            val_cell(row, 1, month_key,              bg=alt.fgColor.rgb if hasattr(alt.fgColor, 'rgb') else 'FFFFFF', halign='left')
            val_cell(row, 2, v['policy_count'],      halign='center')
            val_cell(row, 3, v['member_count'],      halign='center')
            val_cell(row, 4, round(v['gross_premium'], 0), num_fmt='#,##0')
            val_cell(row, 5, round(v['grand_total'],  0),  num_fmt='#,##0')
            sg += v['gross_premium']; st += v['grand_total']
            sc += v['policy_count']; sm += v['member_count']
            row += 1
        for c, v in enumerate([('TOTAL', sc, sm, round(sg,0), round(st,0))][0], 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = Font(name='Raleway', bold=True, size=9.5, color='FFFFFF')
            cell.fill = _PF('solid', fgColor=PRI)
            cell.alignment = Alignment(horizontal='right' if c > 1 else 'left', vertical='center')
            if c >= 4:
                cell.number_format = '#,##0'
        row += 2

        # ── Broker performance ──────────────────────────────────────────────
        brokers = defaultdict(lambda: {'count': 0, 'members': 0, 'gross': 0.0, 'grand': 0.0})
        for r in br_rows:
            b = r.get('broker') or 'Unknown'
            brokers[b]['count']   += 1
            brokers[b]['members'] += int(r.get('member_count') or 0)
            brokers[b]['gross']   += float(r.get('total_net') or 0) + float(r.get('total_maternity') or 0)
            brokers[b]['grand']   += float(r.get('grand_total') or 0)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value='Broker Performance').font = Font(name='Raleway', bold=True, size=11, color=PRI)
        row += 1
        for c, h in enumerate(['Broker', 'Policies', 'Members', 'Gross Premium', 'Grand Total'], 1):
            hdr_cell(row, c, h)
        row += 1
        for bname, v in sorted(brokers.items(), key=lambda x: -x[1]['grand']):
            val_cell(row, 1, bname, halign='left')
            val_cell(row, 2, v['count'],  halign='center')
            val_cell(row, 3, v['members'], halign='center')
            val_cell(row, 4, round(v['gross'], 0), num_fmt='#,##0')
            val_cell(row, 5, round(v['grand'], 0), num_fmt='#,##0')
            row += 1
        row += 1

        # ── Take Rate by Plan ───────────────────────────────────────────────
        tr_buckets = defaultdict(lambda: {'premium': 0.0, 'commission': 0.0, 'count': 0})
        for r in tr_rows:
            plan  = r.get('plan') or 'Unknown'
            gross = float(r.get('total_net') or 0) + float(r.get('total_maternity') or 0)
            cpct  = sum(float(r.get(k, 0) or 0) for k in ('rm_broker','rm_insurer','rm_wellx','rm_tpa','rm_insurance_tax'))
            tr_buckets[plan]['premium']    += gross
            tr_buckets[plan]['commission'] += gross * cpct / 100.0
            tr_buckets[plan]['count']      += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value='Take Rate by Plan').font = Font(name='Raleway', bold=True, size=11, color=PRI)
        row += 1
        for c, h in enumerate(['Plan', 'Policies', 'Total Premium', 'Commission Amount', 'Take Rate %'], 1):
            hdr_cell(row, c, h, bg=ORG)
        row += 1
        tp = tc = 0.0
        for plan, v in sorted(tr_buckets.items()):
            prem = round(v['premium'], 0); comm = round(v['commission'], 0)
            rate = round(v['commission'] / v['premium'] * 100, 2) if v['premium'] else 0
            val_cell(row, 1, plan, halign='left')
            val_cell(row, 2, v['count'], halign='center')
            val_cell(row, 3, prem, num_fmt='#,##0')
            val_cell(row, 4, comm, num_fmt='#,##0')
            val_cell(row, 5, rate, num_fmt='0.00"%"')
            tp += v['premium']; tc += v['commission']
            row += 1
        total_rate = round(tc / tp * 100, 2) if tp else 0
        for c, v in enumerate([('TOTAL', '', round(tp,0), round(tc,0), total_rate)][0], 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = Font(name='Raleway', bold=True, size=9.5, color='FFFFFF')
            cell.fill = _PF('solid', fgColor=ORG)
            cell.alignment = Alignment(horizontal='right' if c > 1 else 'left', vertical='center')
            if c in (3, 4):
                cell.number_format = '#,##0'
            elif c == 5:
                cell.number_format = '0.00"%"'
        row += 2

        # ── Powered by footer ───────────────────────────────────────────────
        _xl_powered_row(ws, row, last_col=5)
        _xl_add_logo(ws, f'F{row}', _POWERED_LOGO_PATH, w=110, h=28)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"Wellx Policies Report {today_str}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Policy List Excel Export (filtered) ───────────────────────────────────────
@app.route('/api/policies/export')
def api_policies_export():
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'Database not configured'}), 503
    try:
        from openpyxl import Workbook as _WB
        from openpyxl.styles import PatternFill as _PF

        # Reuse same filter logic as /api/policies
        q = supa.table('policies').select(
            'id,created_at,company_name,broker,rm_person,plan,plan_type,'
            'member_count,total_net,total_maternity,grand_total,recon_match'
        )
        search    = request.args.get('search', '').strip().lower()
        broker    = request.args.get('broker', '').strip()
        rm        = request.args.get('rm', '').strip()
        plan      = request.args.get('plan', '').strip()
        plan_type = request.args.get('plan_type', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to   = request.args.get('date_to', '').strip()
        if broker:
            q = q.eq('broker', broker)
        if rm:
            q = q.eq('rm_person', rm)
        if plan:
            q = q.eq('plan', plan)
        if plan_type:
            q = q.eq('plan_type', plan_type)
        if date_from:
            q = q.gte('created_at', date_from)
        if date_to:
            q = q.lte('created_at', date_to + 'T23:59:59')
        rows = q.order('created_at', desc=True).execute().data or []
        if search:
            rows = [r for r in rows
                    if search in (r.get('company_name') or '').lower()
                    or search in (r.get('broker') or '').lower()
                    or search in (r.get('rm_person') or '').lower()]

        today_str = datetime.now().strftime('%d %b %Y')
        wb = _WB()
        ws = wb.active
        ws.title = 'Policies'
        PRI = '003780'; ORG = 'E86F2C'; LGT = 'EFF4FB'

        col_widths = [6, 14, 30, 20, 18, 14, 12, 10, 18, 10]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = _make_report_header(ws,
            f'Wellx Policies List — as of {today_str}',
            subtitle=f'{len(rows)} policies · Filtered export',
            col_count=10) + 1

        hdrs = ['#', 'Date Saved', 'Company', 'Broker', 'RM', 'Plan', 'Type', 'Members', 'Total Premium', 'Match']
        for c, h in enumerate(hdrs, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font      = Font(name='Inter', bold=True, size=9, color='FFFFFF')
            cell.fill      = _PF('solid', fgColor=PRI)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        for i, r in enumerate(rows, 1):
            bg = LGT if i % 2 == 0 else 'FFFFFF'
            gross = float(r.get('total_net') or 0) + float(r.get('total_maternity') or 0)
            match_txt = '✅ Match' if r.get('recon_match') is True else ('❌ Mismatch' if r.get('recon_match') is False else '—')
            dt_str = (r.get('created_at') or '')[:10]
            try:
                dt_val = datetime.strptime(dt_str, '%Y-%m-%d').strftime('%d %b %Y')
            except Exception:
                dt_val = dt_str
            vals = [i, dt_val, r.get('company_name',''), r.get('broker',''),
                    r.get('rm_person',''), r.get('plan',''), r.get('plan_type',''),
                    int(r.get('member_count') or 0), round(gross, 0), match_txt]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font      = Font(name='Inter', size=9, color='1e293b')
                cell.fill      = _PF('solid', fgColor=bg)
                cell.alignment = Alignment(
                    horizontal='center' if c in (1,6,7,8,10) else ('right' if c == 9 else 'left'),
                    vertical='center')
                if c == 9:
                    cell.number_format = '#,##0'
            row += 1

        row += 1
        _xl_powered_row(ws, row, last_col=10)
        _xl_add_logo(ws, f'I{row}', _POWERED_LOGO_PATH, w=110, h=26)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        fname = f"Wellx Policies {today_str}.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── DB diagnostics ────────────────────────────────────────────────────────────

@app.route('/api/db_check')
def api_db_check():
    supa = get_supa()
    connected = False
    err = ''
    if supa:
        try:
            supa.table('policies').select('id').limit(1).execute()
            connected = True
        except Exception as e:
            err = str(e)
    return jsonify({
        'supabase_available':    _DB_AVAILABLE,
        'SUPABASE_URL_set':      bool(os.environ.get('SUPABASE_URL')),
        'SUPABASE_ANON_KEY_set': bool(os.environ.get('SUPABASE_ANON_KEY')),
        'connected':             connected,
        'error':                 err,
    })


# ── Filter options for policies page ─────────────────────────────────────────

@app.route('/api/policies/meta')
def api_policies_meta():
    supa = get_supa()
    if not supa:
        return jsonify({'brokers': [], 'rms': [], 'plans': [], 'plan_types': []})
    try:
        res = supa.table('policies').select('broker,rm_person,plan,plan_type').execute()
        rows = res.data or []
        return jsonify({
            'brokers':    sorted({r['broker']    for r in rows if r.get('broker')}),
            'rms':        sorted({r['rm_person'] for r in rows if r.get('rm_person')}),
            'plans':      sorted({r['plan']      for r in rows if r.get('plan')}),
            'plan_types': sorted({r['plan_type'] for r in rows if r.get('plan_type')}),
        })
    except Exception as e:
        return jsonify({'brokers': [], 'rms': [], 'plans': [], 'plan_types': []})


# ── Settings: Underwriters / RM Persons / Brokers ────────────────────────────

@app.route('/settings')
def page_settings():
    return render_template('settings.html')


def _people_list(table_name):
    supa = get_supa()
    if not supa:
        return jsonify([])
    try:
        res = supa.table(table_name).select('*').order('name').execute()
        return jsonify(res.data or [])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _people_add(table_name):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'DB not configured'}), 503
    data = request.get_json(force=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400
    try:
        res = supa.table(table_name).insert({'name': name}).execute()
        return jsonify(res.data[0] if res.data else {}), 201
    except Exception as e:
        msg = str(e)
        if 'unique' in msg.lower() or 'duplicate' in msg.lower():
            return jsonify({'error': 'Already exists'}), 409
        return jsonify({'error': msg}), 500


def _people_edit(table_name, rid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'DB not configured'}), 503
    data = request.get_json(force=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name required'}), 400
    try:
        res = supa.table(table_name).update({'name': name}).eq('id', rid).execute()
        return jsonify(res.data[0] if res.data else {})
    except Exception as e:
        msg = str(e)
        if 'unique' in msg.lower() or 'duplicate' in msg.lower():
            return jsonify({'error': 'Already exists'}), 409
        return jsonify({'error': msg}), 500


# policy column that links to each people table
_PEOPLE_POLICY_COL = {
    'underwriters': 'underwriter',
    'rm_persons':   'rm_person',
    'brokers':      'broker',
}

def _people_delete(table_name, rid):
    supa = get_supa()
    if not supa:
        return jsonify({'error': 'DB not configured'}), 503
    try:
        # Check if any policy references this person
        row = supa.table(table_name).select('name').eq('id', rid).single().execute()
        name = row.data['name'] if row.data else None
        if name:
            pol_col = _PEOPLE_POLICY_COL.get(table_name)
            if pol_col:
                check = supa.table('policies').select('id', count='exact').eq(pol_col, name).limit(1).execute()
                if (check.count or 0) > 0:
                    return jsonify({'error': f'Cannot delete — {check.count} policy record(s) use this name.'}), 409
        supa.table(table_name).delete().eq('id', rid).execute()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Underwriters
@app.route('/api/settings/underwriters', methods=['GET'])
def api_underwriters_list():
    return _people_list('underwriters')

@app.route('/api/settings/underwriters', methods=['POST'])
def api_underwriters_add():
    return _people_add('underwriters')

@app.route('/api/settings/underwriters/<int:rid>', methods=['PUT'])
def api_underwriters_edit(rid):
    return _people_edit('underwriters', rid)

@app.route('/api/settings/underwriters/<int:rid>', methods=['DELETE'])
def api_underwriters_delete(rid):
    return _people_delete('underwriters', rid)


# RM Persons
@app.route('/api/settings/rm_persons', methods=['GET'])
def api_rm_list():
    return _people_list('rm_persons')

@app.route('/api/settings/rm_persons', methods=['POST'])
def api_rm_add():
    return _people_add('rm_persons')

@app.route('/api/settings/rm_persons/<int:rid>', methods=['PUT'])
def api_rm_edit(rid):
    return _people_edit('rm_persons', rid)

@app.route('/api/settings/rm_persons/<int:rid>', methods=['DELETE'])
def api_rm_delete(rid):
    return _people_delete('rm_persons', rid)


# Brokers
@app.route('/api/settings/brokers', methods=['GET'])
def api_brokers_list():
    return _people_list('brokers')

@app.route('/api/settings/brokers', methods=['POST'])
def api_brokers_add():
    return _people_add('brokers')

@app.route('/api/settings/brokers/<int:rid>', methods=['PUT'])
def api_brokers_edit(rid):
    return _people_edit('brokers', rid)

@app.route('/api/settings/brokers/<int:rid>', methods=['DELETE'])
def api_brokers_delete(rid):
    return _people_delete('brokers', rid)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    app.run(debug=False, port=port, host='0.0.0.0')
